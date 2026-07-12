from pathlib import Path

from openpyxl import load_workbook

from app.db import db
from app.repositories.panel_repository import (
    build_internal_name,
    normalize_ip,
    normalize_mac,
)


EXCEL_PATH = Path("панели123.xlsx")


def clean(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Файл не найден: {EXCEL_PATH.resolve()}"
        )

    workbook = load_workbook(
        EXCEL_PATH,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    rows_to_insert: list[tuple[str, str, str, str, str]] = []
    mac_addresses: set[str] = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        address = clean(row[0])
        entrance = clean(row[1])
        mac = normalize_mac(clean(row[2]))
        ip = normalize_ip(clean(row[3]))

        if not address and not entrance and not mac and not ip:
            continue

        if not address:
            raise ValueError(
                f"Строка {row_number}: не указан адрес"
            )

        if not mac:
            raise ValueError(
                f"Строка {row_number}: не указан MAC-адрес"
            )

        if mac in mac_addresses:
            raise ValueError(
                f"Строка {row_number}: повторяется MAC {mac}"
            )

        mac_addresses.add(mac)

        internal_name = build_internal_name(
            address,
            entrance,
        )

        rows_to_insert.append(
            (
                address,
                entrance,
                internal_name,
                mac,
                ip,
            )
        )

    if not rows_to_insert:
        raise ValueError(
            "В Excel-файле не найдено ни одной панели"
        )

    print(f"Найдено панелей в Excel: {len(rows_to_insert)}")
    print("Текущие панели будут полностью удалены.")

    confirmation = input(
        'Для продолжения введи слово "ЗАМЕНИТЬ": '
    ).strip()

    if confirmation != "ЗАМЕНИТЬ":
        print("Операция отменена.")
        return

    with db() as conn:
        # Удаляем привязки панелей к управляющим компаниям.
        conn.execute(
            """
            DELETE FROM uk_group_panels
            """
        )

        # Полностью удаляем старые панели.
        conn.execute(
            """
            DELETE FROM panels
            """
        )

        # Сбрасываем счётчик ID.
        conn.execute(
            """
            DELETE FROM sqlite_sequence
            WHERE name = 'panels'
            """
        )

        conn.executemany(
            """
            INSERT INTO panels(
                address,
                entrance,
                name,
                mac,
                ip,
                tags,
                enabled
            )
            VALUES (?, ?, ?, ?, ?, '', 1)
            """,
            rows_to_insert,
        )

    print(
        f"Готово. Загружено панелей: {len(rows_to_insert)}"
    )


if __name__ == "__main__":
    main()