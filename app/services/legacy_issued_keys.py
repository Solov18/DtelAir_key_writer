"""Import previously issued physical keys without programming any panel.

The source workbook is an accounting migration. Existing physical keys are
reused, missing keys are created locally, and resident assignments are linked
to addresses that already exist in the panel register. This module deliberately
does not import the panel API client or the external CRM writer.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from app.db import db
from app.repositories.key_repository import set_key_assignment_on_connection
from app.services.keys import normalize_hex_value
from app.services.parser import (
    APARTMENT_RE,
    _select_detected_address,
    find_address_candidates,
    get_panel_addresses,
    remove_noise,
    split_address_tokens,
)


LEGACY_ASSIGNMENT_NOTE = "Перенос ранее выданного ключа; без записи на панели"
LEGACY_FREE_NOTE = "Перенос свободного ключа из старой базы; без записи на панели"

DEFAULT_TYPE_ALIASES = {
    "синий": "Синий",
    "синие": "Синий",
    "премиальный": "Премиальные",
    "премиальные": "Премиальные",
    "уникальный": "Уникальные",
    "уникальные": "Уникальные",
    "стикер": "Стикер",
    "стикеры": "Стикер",
    "розовый": "Розовый",
    "розовые": "Розовый",
}

_HEX_RE = re.compile(r"[0-9A-F]{6,16}")


@dataclass(frozen=True)
class LegacyIssuedKeyRow:
    source: str
    type_name: str
    number: str
    hex_value: str
    assignment: str


def _cell_text(value) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def read_legacy_issued_workbook(path: str | Path) -> list[LegacyIssuedKeyRow]:
    """Read a normalized XLSX copy of the old register."""

    workbook = load_workbook(
        filename=Path(path),
        read_only=True,
        data_only=True,
    )
    result: list[LegacyIssuedKeyRow] = []
    try:
        for worksheet in workbook.worksheets:
            current_type = ""
            for line_number, values in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                first = _cell_text(values[0] if len(values) > 0 else "")
                second = _cell_text(values[1] if len(values) > 1 else "")
                third = _cell_text(values[2] if len(values) > 2 else "")
                if not first and not second and not third:
                    continue
                if first and not second and not third and not first.isdigit():
                    current_type = first
                    continue
                if not first:
                    continue
                result.append(
                    LegacyIssuedKeyRow(
                        source=f"Лист «{worksheet.title}», строка {line_number}",
                        type_name=current_type,
                        number=first,
                        hex_value=second,
                        assignment=third,
                    )
                )
    finally:
        workbook.close()
    return result


def _type_key(value: str) -> str:
    return re.sub(
        r"\s+",
        " ",
        str(value or "").strip().casefold().replace("ё", "е"),
    )


def _address_key(value: str) -> tuple[tuple[str, ...], str]:
    parsed = split_address_tokens(remove_noise(value))
    street = tuple(token for token in parsed["street_tokens"] if token != "дом")
    house = parsed["house"]
    extras = [token for token in parsed["extra_tokens"] if token != "дом"]
    if extras and len(extras[0]) == 1 and extras[0].isalpha():
        suffix = extras[0]
        if house.endswith("л"):
            house = f"{house[:-1]}{suffix}"
        elif re.fullmatch(r"\d+", house):
            house = f"{house}{suffix}"
    return street, house


def _address_catalog_index(catalog: list[dict]) -> dict[tuple, list[dict]]:
    indexed: dict[tuple, list[dict]] = defaultdict(list)
    for item in catalog:
        street = tuple(token for token in item["street_tokens"] if token != "дом")
        indexed[(street, item["house"])].append(item)
    return indexed


def _resolve_address(
    raw_assignment: str,
    catalog: list[dict],
    indexed_catalog: dict[tuple, list[dict]],
) -> tuple[str, str, str, list[dict]]:
    apartment_matches = list(APARTMENT_RE.finditer(raw_assignment or ""))
    if len(apartment_matches) > 1:
        return "", "", "multiple_addresses", []
    apartment = apartment_matches[0].group(1) if apartment_matches else ""

    exact = indexed_catalog.get(_address_key(raw_assignment), [])
    if len(exact) == 1:
        return exact[0]["address"], apartment, "exact", exact
    if len(exact) > 1:
        return "", apartment, "ambiguous", exact

    candidates = find_address_candidates(
        raw_assignment,
        limit=3,
        address_catalog=catalog,
    )
    address, status = _select_detected_address(candidates)
    best = candidates[0] if candidates else {}
    if (
        status in {"exact", "similar"}
        and best.get("house_match") == "house_exact"
        and float(best.get("street_confidence") or 0) >= 0.80
    ):
        return address, apartment, status, candidates
    return "", apartment, status, candidates


def _initial_report(*, dry_run: bool) -> dict:
    return {
        "dry_run": bool(dry_run),
        "mode": LEGACY_ASSIGNMENT_NOTE,
        "source_rows": 0,
        "rows_with_address": 0,
        "found_keys": 0,
        "created_only_crm": 0,
        "linked_to_addresses": 0,
        "assignments_created_or_changed": 0,
        "already_linked": 0,
        "addresses_not_found": 0,
        "address_matches_exact": 0,
        "address_matches_similar": 0,
        "skipped_without_address": 0,
        "input_duplicates": 0,
        "errors": 0,
        "panel_requests": 0,
        "error_details": [],
        "address_details": [],
        "similar_match_details": [],
    }


def _append_detail(report: dict, field: str, text: str, limit: int = 200) -> None:
    if len(report[field]) < limit:
        report[field].append(text)


def import_legacy_issued_keys(
    rows: Iterable[LegacyIssuedKeyRow],
    *,
    actor: str,
    dry_run: bool = True,
    source_name: str = "",
    source_hash: str = "",
) -> dict:
    """Plan or apply an idempotent legacy-issued-key migration.

    The apply path changes only ``keys``, ``key_assignments`` and one summary
    row in ``operation_log``. ``dry_run=True`` performs no database writes.
    """

    source_rows = list(rows)
    report = _initial_report(dry_run=dry_run)
    report["source_rows"] = len(source_rows)

    usable: list[tuple[LegacyIssuedKeyRow, str]] = []
    by_hex: dict[str, list[LegacyIssuedKeyRow]] = defaultdict(list)
    by_number: dict[tuple[str, str], list[LegacyIssuedKeyRow]] = defaultdict(list)
    for row in source_rows:
        if not row.assignment.strip():
            report["skipped_without_address"] += 1
            continue
        report["rows_with_address"] += 1
        clean_hex = normalize_hex_value(row.hex_value)
        if not _HEX_RE.fullmatch(clean_hex):
            report["errors"] += 1
            _append_detail(
                report,
                "error_details",
                f"{row.source}: отсутствует или некорректен HEX «{row.hex_value}».",
            )
            continue
        usable.append((row, clean_hex))
        by_hex[clean_hex].append(row)
        by_number[(_type_key(row.type_name), row.number.casefold())].append(row)

    conflicting_hex = {
        value
        for value, grouped in by_hex.items()
        if len(
            {
                (
                    _type_key(item.type_name),
                    item.number.casefold(),
                    _type_key(item.assignment),
                )
                for item in grouped
            }
        )
        > 1
    }
    conflicting_numbers = {
        value
        for value, grouped in by_number.items()
        if len({normalize_hex_value(item.hex_value) for item in grouped}) > 1
    }
    for clean_hex in sorted(conflicting_hex):
        report["errors"] += 1
        _append_detail(
            report,
            "error_details",
            f"HEX {clean_hex}: в исходном файле указаны разные номера или назначения.",
        )
    for type_number in sorted(conflicting_numbers):
        report["errors"] += 1
        _append_detail(
            report,
            "error_details",
            f"{type_number[0]} №{type_number[1]}: в исходном файле указаны разные HEX.",
        )

    catalog = get_panel_addresses()
    indexed_catalog = _address_catalog_index(catalog)
    seen_fingerprints: set[tuple] = set()

    with db() as conn:
        type_rows = [dict(row) for row in conn.execute("SELECT id, name FROM key_types")]
        types_by_key = {_type_key(row["name"]): row for row in type_rows}
        for alias, canonical in DEFAULT_TYPE_ALIASES.items():
            canonical_row = types_by_key.get(_type_key(canonical))
            if canonical_row:
                types_by_key[_type_key(alias)] = canonical_row

        key_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT k.*, kt.name AS type_name
                FROM keys k
                JOIN key_types kt ON kt.id = k.key_type_id
                """
            )
        ]
        keys_by_hex = {
            normalize_hex_value(row["hex_value"]): row
            for row in key_rows
            if normalize_hex_value(row["hex_value"])
        }
        keys_by_number = {
            (int(row["key_type_id"]), str(row["number"]).casefold()): row
            for row in key_rows
        }
        active_assignments = {
            int(row["key_id"]): dict(row)
            for row in conn.execute("SELECT * FROM key_assignments WHERE active = 1")
        }

        for row, clean_hex in usable:
            type_number_key = (_type_key(row.type_name), row.number.casefold())
            if clean_hex in conflicting_hex or type_number_key in conflicting_numbers:
                continue

            fingerprint = (
                clean_hex,
                _type_key(row.type_name),
                row.number.casefold(),
                _type_key(row.assignment),
            )
            if fingerprint in seen_fingerprints:
                report["input_duplicates"] += 1
                continue
            seen_fingerprints.add(fingerprint)

            key_type = types_by_key.get(_type_key(row.type_name))
            if not key_type:
                report["errors"] += 1
                _append_detail(
                    report,
                    "error_details",
                    f"{row.source}: тип ключа «{row.type_name}» отсутствует в CRM.",
                )
                continue

            address, apartment, address_status, candidates = _resolve_address(
                row.assignment,
                catalog,
                indexed_catalog,
            )
            if not address:
                report["addresses_not_found"] += 1
                suggestions = ", ".join(item["address"] for item in candidates[:3])
                suffix = f"; варианты: {suggestions}" if suggestions else ""
                _append_detail(
                    report,
                    "address_details",
                    f"{row.source}: «{row.assignment}» не сопоставлен{suffix}.",
                )
                continue
            if address_status == "exact":
                report["address_matches_exact"] += 1
            else:
                report["address_matches_similar"] += 1
                confidence = candidates[0].get("confidence") if candidates else None
                _append_detail(
                    report,
                    "similar_match_details",
                    f"{row.source}: «{row.assignment}» → «{address}» "
                    f"(уверенность {confidence}).",
                    limit=1000,
                )

            number_key = (int(key_type["id"]), row.number.casefold())
            by_hex_key = keys_by_hex.get(clean_hex)
            by_number_key = keys_by_number.get(number_key)
            if (
                by_hex_key
                and by_number_key
                and int(by_hex_key["id"]) != int(by_number_key["id"])
            ):
                report["errors"] += 1
                _append_detail(
                    report,
                    "error_details",
                    f"{row.source}: HEX и номер принадлежат разным ключам CRM.",
                )
                continue

            existing = by_hex_key or by_number_key
            if existing:
                if (
                    int(existing["key_type_id"]) != int(key_type["id"])
                    or str(existing["number"]).casefold() != row.number.casefold()
                ):
                    report["errors"] += 1
                    _append_detail(
                        report,
                        "error_details",
                        f"{row.source}: HEX уже сохранён как {existing['type_name']} №{existing['number']}.",
                    )
                    continue
                existing_hex = normalize_hex_value(existing["hex_value"])
                if existing_hex and existing_hex != clean_hex:
                    report["errors"] += 1
                    _append_detail(
                        report,
                        "error_details",
                        f"{row.source}: у ключа №{row.number} в CRM сохранён другой HEX.",
                    )
                    continue
                report["found_keys"] += 1
                key_id = int(existing["id"])
                if not dry_run and not existing_hex:
                    conn.execute(
                        "UPDATE keys SET hex_value = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                        (clean_hex, key_id),
                    )
                    existing["hex_value"] = clean_hex
                    keys_by_hex[clean_hex] = existing
            else:
                report["created_only_crm"] += 1
                if dry_run:
                    key_id = -report["created_only_crm"]
                    existing = {
                        "id": key_id,
                        "key_type_id": int(key_type["id"]),
                        "number": row.number,
                        "hex_value": clean_hex,
                        "type_name": key_type["name"],
                    }
                else:
                    cursor = conn.execute(
                        """
                        INSERT INTO keys(
                            key_type_id, number, hex_value, key_type,
                            status, note, is_used, created_by
                        )
                        VALUES (?, ?, ?, ?, 'issued_resident', ?, 1, ?)
                        """,
                        (
                            key_type["id"],
                            row.number,
                            clean_hex,
                            key_type["name"],
                            LEGACY_ASSIGNMENT_NOTE,
                            actor,
                        ),
                    )
                    key_id = int(cursor.lastrowid)
                    existing = {
                        "id": key_id,
                        "key_type_id": int(key_type["id"]),
                        "number": row.number,
                        "hex_value": clean_hex,
                        "type_name": key_type["name"],
                    }
                keys_by_hex[clean_hex] = existing
                keys_by_number[number_key] = existing

            current = active_assignments.get(key_id)
            same_assignment = bool(
                current
                and current["assignment_type"] == "resident"
                and str(current.get("address") or "").strip().casefold()
                == address.strip().casefold()
                and str(current.get("apartment") or "").strip().casefold()
                == apartment.strip().casefold()
            )
            if same_assignment:
                report["already_linked"] += 1
            else:
                report["assignments_created_or_changed"] += 1

            if not dry_run:
                set_key_assignment_on_connection(
                    conn,
                    key_id,
                    "resident",
                    address=address,
                    apartment=apartment,
                    assigned_by=actor,
                    note=LEGACY_ASSIGNMENT_NOTE,
                )
            active_assignments[key_id] = {
                "key_id": key_id,
                "assignment_type": "resident",
                "address": address,
                "apartment": apartment,
            }
            report["linked_to_addresses"] += 1

        if not dry_run:
            status = (
                "success"
                if report["errors"] == 0 and report["addresses_not_found"] == 0
                else "warning"
            )
            details = (
                f"Источник: {source_name or 'файл'}; SHA-256: {source_hash or '-'}; "
                f"найдено: {report['found_keys']}; создано: {report['created_only_crm']}; "
                f"привязано: {report['linked_to_addresses']}; "
                f"адреса не найдены: {report['addresses_not_found']}; "
                f"ошибки: {report['errors']}. Запросы к панелям не выполнялись."
            )
            conn.execute(
                """
                INSERT INTO operation_log(
                    mode, action, object_type, object_name, details,
                    status, hex_value, mac,
                    username, user_full_name, comment
                )
                VALUES (?, 'legacy_issued_import', 'Ключ', ?, ?, ?, '-', '', ?, ?, ?)
                """,
                (
                    "legacy_issued_import",
                    source_name or "Перенос ранее выданных ключей",
                    details,
                    status,
                    actor,
                    actor,
                    LEGACY_ASSIGNMENT_NOTE,
                ),
            )

    return report


def import_legacy_free_keys(
    rows: Iterable[LegacyIssuedKeyRow],
    *,
    actor: str,
    dry_run: bool = True,
    source_name: str = "",
    source_hash: str = "",
) -> dict:
    """Create only missing, unassigned keys from address-free source rows.

    Existing keys are treated as immutable: their status, HEX and assignments
    are never changed. The apply path writes only new rows to ``keys`` and one
    summary row to ``operation_log``. No panel or external CRM client is used.
    """

    source_rows = list(rows)
    free_rows = [row for row in source_rows if not row.assignment.strip()]
    report = {
        "dry_run": bool(dry_run),
        "mode": LEGACY_FREE_NOTE,
        "source_rows": len(source_rows),
        "free_rows_in_file": len(free_rows),
        "created_in_crm": 0,
        "already_existed": 0,
        "unprocessed": 0,
        "input_duplicates": 0,
        "panel_requests": 0,
        "error_details": [],
    }

    by_hex: dict[str, list[LegacyIssuedKeyRow]] = defaultdict(list)
    by_number: dict[tuple[str, str], list[LegacyIssuedKeyRow]] = defaultdict(list)
    normalized_rows: list[tuple[LegacyIssuedKeyRow, str]] = []
    for row in free_rows:
        clean_hex = normalize_hex_value(row.hex_value)
        if not _HEX_RE.fullmatch(clean_hex):
            report["unprocessed"] += 1
            _append_detail(
                report,
                "error_details",
                f"{row.source}: отсутствует или некорректен HEX «{row.hex_value}».",
                limit=1000,
            )
            continue
        normalized_rows.append((row, clean_hex))
        by_hex[clean_hex].append(row)
        by_number[(_type_key(row.type_name), row.number.casefold())].append(row)

    conflicting_hex = {
        value
        for value, grouped in by_hex.items()
        if len(
            {
                (_type_key(item.type_name), item.number.casefold())
                for item in grouped
            }
        ) > 1
    }
    conflicting_numbers = {
        value
        for value, grouped in by_number.items()
        if len({normalize_hex_value(item.hex_value) for item in grouped}) > 1
    }
    seen_fingerprints: set[tuple[str, str, str]] = set()

    with db() as conn:
        type_rows = [dict(row) for row in conn.execute("SELECT id, name FROM key_types")]
        types_by_key = {_type_key(row["name"]): row for row in type_rows}
        for alias, canonical in DEFAULT_TYPE_ALIASES.items():
            canonical_row = types_by_key.get(_type_key(canonical))
            if canonical_row:
                types_by_key[_type_key(alias)] = canonical_row

        key_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT k.*, kt.name AS type_name
                FROM keys k
                JOIN key_types kt ON kt.id = k.key_type_id
                """
            )
        ]
        keys_by_hex = {
            normalize_hex_value(row["hex_value"]): row
            for row in key_rows
            if normalize_hex_value(row["hex_value"])
        }
        keys_by_number = {
            (int(row["key_type_id"]), str(row["number"]).casefold()): row
            for row in key_rows
        }

        for row, clean_hex in normalized_rows:
            source_type_number = (_type_key(row.type_name), row.number.casefold())
            fingerprint = (clean_hex, *source_type_number)
            if fingerprint in seen_fingerprints:
                report["input_duplicates"] += 1
                continue
            seen_fingerprints.add(fingerprint)

            if clean_hex in conflicting_hex or source_type_number in conflicting_numbers:
                report["unprocessed"] += 1
                _append_detail(
                    report,
                    "error_details",
                    f"{row.source}: конфликт номера или HEX внутри исходного файла.",
                    limit=1000,
                )
                continue

            key_type = types_by_key.get(_type_key(row.type_name))
            if not key_type:
                report["unprocessed"] += 1
                _append_detail(
                    report,
                    "error_details",
                    f"{row.source}: тип ключа «{row.type_name}» отсутствует в CRM.",
                    limit=1000,
                )
                continue

            number_key = (int(key_type["id"]), row.number.casefold())
            by_hex_key = keys_by_hex.get(clean_hex)
            by_number_key = keys_by_number.get(number_key)
            if (
                by_hex_key
                and by_number_key
                and int(by_hex_key["id"]) != int(by_number_key["id"])
            ):
                report["unprocessed"] += 1
                _append_detail(
                    report,
                    "error_details",
                    f"{row.source}: HEX и номер принадлежат разным ключам CRM.",
                    limit=1000,
                )
                continue

            existing = by_hex_key or by_number_key
            if existing:
                if (
                    int(existing["key_type_id"]) != int(key_type["id"])
                    or str(existing["number"]).casefold() != row.number.casefold()
                    or normalize_hex_value(existing["hex_value"]) != clean_hex
                ):
                    report["unprocessed"] += 1
                    _append_detail(
                        report,
                        "error_details",
                        f"{row.source}: номер или HEX конфликтует с ключом CRM "
                        f"{existing['type_name']} №{existing['number']}.",
                        limit=1000,
                    )
                    continue
                report["already_existed"] += 1
                continue

            report["created_in_crm"] += 1
            if dry_run:
                key_id = -report["created_in_crm"]
            else:
                key_id = int(
                    conn.execute(
                        """
                        INSERT INTO keys(
                            key_type_id, number, hex_value, key_type,
                            status, note, is_used, created_by
                        )
                        VALUES (?, ?, ?, ?, 'free', ?, 0, ?)
                        """,
                        (
                            key_type["id"],
                            row.number,
                            clean_hex,
                            key_type["name"],
                            LEGACY_FREE_NOTE,
                            actor,
                        ),
                    ).lastrowid
                )
            created = {
                "id": key_id,
                "key_type_id": int(key_type["id"]),
                "number": row.number,
                "hex_value": clean_hex,
                "type_name": key_type["name"],
            }
            keys_by_hex[clean_hex] = created
            keys_by_number[number_key] = created

        if not dry_run:
            status = "success" if report["unprocessed"] == 0 else "warning"
            details = (
                f"Источник: {source_name or 'файл'}; SHA-256: {source_hash or '-'}; "
                f"строк без адреса: {report['free_rows_in_file']}; "
                f"создано: {report['created_in_crm']}; "
                f"уже существовало: {report['already_existed']}; "
                f"не обработано: {report['unprocessed']}. "
                "Назначения не изменялись. Запросы к панелям не выполнялись."
            )
            conn.execute(
                """
                INSERT INTO operation_log(
                    mode, action, object_type, object_name, details,
                    status, hex_value, mac,
                    username, user_full_name, comment
                )
                VALUES (?, 'legacy_free_import', 'Ключ', ?, ?, ?, '-', '', ?, ?, ?)
                """,
                (
                    "legacy_free_import",
                    source_name or "Перенос свободных ключей",
                    details,
                    status,
                    actor,
                    actor,
                    LEGACY_FREE_NOTE,
                ),
            )

    return report
