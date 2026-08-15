import csv
import io
import re

from openpyxl import load_workbook

from app.db import db
from app.services.keys import normalize_hex_value
from app.services.panels import split_panel_address


KEY_NUMBER_COLUMNS = (
    "номер",
    "№",
    "number",
    "printed_number",
    "напечатанный номер",
)
KEY_HEX_COLUMNS = (
    "hex",
    "hex_value",
    "код",
    "код для вшития",
)
KEY_TYPE_COLUMNS = ("тип", "type", "вид")
KEY_NOTE_COLUMNS = ("комментарий", "comment", "note", "примечание")


def _clean_header(value) -> str:
    return str(value or "").strip().lower()


def _panel_header(value) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", _clean_header(value).replace("ё", "е"))


def _panel_column(headers: list[str], *aliases: str) -> int | None:
    normalized_aliases = {_panel_header(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if _panel_header(header) in normalized_aliases:
            return index
    return None


def _row_text(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return str(row[index] or "").strip()


def _first_value(row: dict, names: tuple[str, ...]):
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip() != "":
            return value
    return ""


def _excel_number(value) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def _read_key_rows(filename: str, content: bytes) -> list[dict]:
    rows: list[dict] = []

    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        lines = text.splitlines()
        delimiter = ";" if lines and ";" in lines[0] else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        for line_number, source in enumerate(reader, start=2):
            item = {_clean_header(key): value for key, value in source.items()}
            item["__sheet_type"] = ""
            item["__source"] = f"CSV, строка {line_number}"
            rows.append(item)
        return rows

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    for worksheet in workbook.worksheets:
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = [_clean_header(value) for value in header_row]

        for line_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            item = {
                headers[index]: values[index]
                for index in range(min(len(headers), len(values)))
                if headers[index]
            }
            item["__sheet_type"] = worksheet.title.strip()
            item["__source"] = f"Лист «{worksheet.title}», строка {line_number}"
            rows.append(item)

    return rows


def import_keys_file(
    filename: str,
    content: bytes,
    created_by: str = "",
) -> dict:
    report = {
        "created_types": 0,
        "added": 0,
        "updated": 0,
        "duplicates": 0,
        "errors": 0,
        "error_details": [],
    }

    try:
        rows = _read_key_rows(filename or "", content)
    except Exception as error:
        report["errors"] = 1
        report["error_details"].append(f"Файл не прочитан: {error}")
        return report

    with db() as conn:
        type_cache = {
            row["name"].strip().lower(): (int(row["id"]), row["name"])
            for row in conn.execute("SELECT id, name FROM key_types")
        }

        # Import files can contain tens of thousands of keys.  Looking up the
        # key number and HEX with two SQL statements for every row made a
        # normal import execute 100,000+ round trips and monopolise a worker
        # for minutes.  Load the two uniqueness maps once, keep them current
        # while parsing the file, and flush writes in batches below.
        existing_keys: dict[tuple[int, str], dict] = {}
        existing_hex: dict[str, dict] = {}
        for stored_row in conn.execute(
            """
            SELECT id, key_type_id, number, hex_value, key_type, note
            FROM keys
            """
        ):
            stored = dict(stored_row.items())
            existing_keys[
                (int(stored["key_type_id"]), str(stored["number"]).lower())
            ] = stored
            stored_hex = str(stored.get("hex_value") or "").upper()
            if stored_hex:
                existing_hex[stored_hex] = stored

        pending_inserts: list[tuple] = []
        pending_updates: list[tuple] = []

        for row in rows:
            source = row.get("__source", "Строка")
            number = _excel_number(_first_value(row, KEY_NUMBER_COLUMNS))
            hex_value = normalize_hex_value(str(_first_value(row, KEY_HEX_COLUMNS)))
            type_name = (
                str(row.get("__sheet_type") or "").strip()
                or str(_first_value(row, KEY_TYPE_COLUMNS)).strip()
                or "Без типа"
            )
            note = str(_first_value(row, KEY_NOTE_COLUMNS)).strip()

            if not number:
                report["errors"] += 1
                report["error_details"].append(f"{source}: не указан номер")
                continue

            if not hex_value:
                report["errors"] += 1
                report["error_details"].append(
                    f"{source}: не указан HEX — ключ не импортирован"
                )
                continue

            if not re.fullmatch(r"[0-9A-F]{6,16}", hex_value):
                report["errors"] += 1
                report["error_details"].append(
                    f"{source}: некорректный HEX «{hex_value}»"
                )
                continue

            type_key = type_name.lower()
            type_data = type_cache.get(type_key)
            if not type_data:
                cursor = conn.execute(
                    """
                    INSERT INTO key_types(name, color, note, enabled)
                    VALUES (?, '#2A9DF4', 'Создано при импорте Excel', 1)
                    """,
                    (type_name,),
                )
                type_data = (int(cursor.lastrowid), type_name)
                type_cache[type_key] = type_data
                report["created_types"] += 1

            key_type_id, canonical_type_name = type_data
            key_identity = (key_type_id, number.lower())
            existing = existing_keys.get(key_identity)
            duplicate_hex = existing_hex.get(hex_value)
            if duplicate_hex is existing:
                duplicate_hex = None

            if duplicate_hex:
                report["duplicates"] += 1
                continue

            if existing:
                current_hex = (existing["hex_value"] or "").upper()
                if current_hex == hex_value or (not hex_value and current_hex):
                    report["duplicates"] += 1
                    continue

                if current_hex and hex_value and current_hex != hex_value:
                    report["errors"] += 1
                    report["error_details"].append(
                        f"{source}: у {canonical_type_name} №{number} уже другой HEX"
                    )
                    continue

                pending_updates.append((hex_value, note, existing["id"]))
                existing["hex_value"] = hex_value or current_hex
                existing["note"] = existing.get("note") or note
                if hex_value:
                    existing_hex[hex_value] = existing
                report["updated"] += 1
                continue

            pending_inserts.append(
                (
                    key_type_id,
                    number,
                    hex_value,
                    canonical_type_name,
                    note,
                    created_by,
                )
            )
            imported = {
                "id": None,
                "key_type_id": key_type_id,
                "number": number,
                "hex_value": hex_value,
                "key_type": canonical_type_name,
                "note": note,
            }
            existing_keys[key_identity] = imported
            existing_hex[hex_value] = imported
            report["added"] += 1

        if pending_updates:
            conn.executemany(
                """
                UPDATE keys
                SET hex_value = CASE WHEN hex_value = '' THEN ? ELSE hex_value END,
                    note = CASE WHEN note = '' THEN ? ELSE note END,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                pending_updates,
            )

        if pending_inserts:
            conn.executemany(
                """
                INSERT INTO keys(
                    key_type_id,
                    number,
                    hex_value,
                    key_type,
                    status,
                    note,
                    created_by
                )
                VALUES (?, ?, ?, ?, 'free', ?, ?)
                """,
                pending_inserts,
            )

    report["error_details"] = report["error_details"][:25]
    return report


def import_panels_csv(content: bytes) -> int:
    text = content.decode("utf-8-sig")
    delimiter = ";" if ";" in text.splitlines()[0] else ","

    reader = csv.DictReader(
        io.StringIO(text),
        delimiter=delimiter,
    )

    values: list[tuple[str, str, str, str, str]] = []

    for row in reader:
        address = (
            row.get("address")
            or row.get("адрес")
            or ""
        ).strip()

        mac = (
            row.get("mac")
            or row.get("MAC")
            or ""
        ).strip().upper()

        name = (
            row.get("name")
            or row.get("название")
            or f"{address} {mac}"
        ).strip()

        entrance = (
            row.get("entrance")
            or row.get("вход")
            or row.get("подъезд")
            or ""
        ).strip()

        tags = (
            row.get("tags")
            or row.get("теги")
            or ""
        ).strip()

        if address and mac:
            values.append((address, entrance, name, mac, tags))

    with db() as conn:
        if values:
            conn.executemany(
                """
                INSERT INTO panels(address, entrance, name, mac, tags)
                VALUES(?,?,?,?,?)
                ON CONFLICT(mac)
                DO UPDATE SET
                    address = excluded.address,
                    entrance = excluded.entrance,
                    name = excluded.name,
                    tags = excluded.tags
                """,
                values,
            )

    return len(values)


def import_panels_excel(filename: str, content: bytes) -> dict:
    result = {
        "added": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
    }

    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        result["errors"] += 1
        return result

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            data_only=True,
            read_only=True,
        )
    except Exception:
        result["errors"] += 1
        return result

    rows = []

    for worksheet in workbook.worksheets:
        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            (),
        )

        headers = [str(header or "").strip() for header in header_row]
        address_index = _panel_column(headers, "Адрес", "Address")
        entrance_index = _panel_column(
            headers,
            "Подъезд / вход",
            "Подъезд",
            "Вход",
            "Entrance",
        )
        ip_index = _panel_column(headers, "IP", "IP-адрес", "IP address")
        mac_index = _panel_column(
            headers,
            "MAC",
            "MAC-адрес",
            "МАС",
            "МАК",
        )
        tags_index = _panel_column(headers, "Теги", "Tags")

        if address_index is None or mac_index is None:
            continue

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            raw_address = _row_text(row, address_index)
            raw_entrance = _row_text(row, entrance_index)
            raw_ip = _row_text(row, ip_index)
            raw_mac = _row_text(row, mac_index).upper()
            raw_tags = _row_text(row, tags_index)

            mac = raw_mac.replace(" ", "")

            if not raw_address or not mac:
                result["skipped"] += 1
                continue

            if not re.fullmatch(r"[0-9A-F]{2}(:[0-9A-F]{2}){5}", mac):
                result["errors"] += 1
                continue

            if entrance_index is None:
                address, entrance = split_panel_address(raw_address)
            else:
                address, entrance = raw_address, raw_entrance
            name = f"{address} {entrance}".strip()

            rows.append(
                {
                    "address": address,
                    "entrance": entrance,
                    "name": name,
                    "mac": mac,
                    "tags": raw_tags,
                    "ip": raw_ip,
                }
            )

    with db() as conn:
        existing_by_mac = {
            str(stored["mac"] or "").upper(): dict(stored.items())
            for stored in conn.execute("SELECT * FROM panels")
            if str(stored["mac"] or "").strip()
        }
        upserts: list[tuple[str, str, str, str, str, str]] = []

        for item in rows:
            existing = existing_by_mac.get(item["mac"])

            if existing:
                old = existing
                address = item["address"] or (old.get("address") or "")
                entrance = item["entrance"] or (old.get("entrance") or "")
                ip = item["ip"] or (old.get("ip") or "")
                tags = item["tags"] or (old.get("tags") or "")
                name = f"{address} {entrance}".strip()

                changed = (
                    old.get("address") != address
                    or old.get("entrance") != entrance
                    or old.get("name") != name
                    or (old.get("ip") or "") != ip
                    or (old.get("tags") or "") != tags
                )

                if changed:
                    result["updated"] += 1
                else:
                    result["skipped"] += 1

            else:
                result["added"] += 1

                address = item["address"]
                entrance = item["entrance"]
                name = item["name"]
                tags = item["tags"]
                ip = item["ip"]

            merged = {
                **(existing or {}),
                "address": address,
                "entrance": entrance,
                "name": name,
                "tags": tags,
                "ip": ip,
                "mac": item["mac"],
            }
            existing_by_mac[item["mac"]] = merged
            upserts.append((address, entrance, name, item["mac"], tags, ip))

        if upserts:
            conn.executemany(
                """
                INSERT INTO panels(address, entrance, name, mac, tags, ip)
                VALUES(?,?,?,?,?,?)
                ON CONFLICT(mac)
                DO UPDATE SET
                    address = excluded.address,
                    entrance = excluded.entrance,
                    name = excluded.name,
                    tags = excluded.tags,
                    ip = excluded.ip
                """,
                upserts,
            )

    return result
