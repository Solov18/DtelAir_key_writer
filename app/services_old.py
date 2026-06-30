import csv, io, json, re
from typing import Iterable
import requests
from openpyxl import load_workbook
from .db import db
from .settings import settings

KEY_RE = re.compile(r'(?<!\d)(\d{4,6})(?!\d)')
PHONE_RE = re.compile(r'(?:\+?7|8)[\s\-()]?\d{3}[\s\-()]?\d{3}[\s\-()]?\d{2}[\s\-()]?\d{2}')
APT_RE = re.compile(r'(?:кв(?:артира)?\.?|квартира)\s*[:№#-]?\s*(\d+)', re.I)
ADDRESS_RE = re.compile(r'((?:[А-Яа-яЁё\w.\-]+\s+){0,3}(?:улица|ул\.?|проспект|пр\.?|пер\.?|переулок|шоссе|проезд)?\s*[А-Яа-яЁё\w.\-]+\s*(?:д\.?|дом)?\s*\d+[А-Яа-яA-Za-z/\-.]*(?:\s*(?:к|корпус|литер)\.?\s*\d+)?(?:\s*/\s*\d+)?)', re.I)


def normalize(s: str) -> str:
    s = (s or '').lower().replace('ё','е')
    s = re.sub(r'\b(улица|ул\.|ул|дом|д\.)\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def parse_message(text: str) -> dict:
    clean = PHONE_RE.sub(' ', text or '')
    apt = APT_RE.search(text or '')
    numbers = []
    for n in KEY_RE.findall(clean):
        if n not in numbers:
            numbers.append(n)
    # address: prefer line with known address words or before apartment
    address = ''
    lines = [x.strip(' ,.;') for x in (text or '').splitlines() if x.strip()]
    for line in lines:
        if re.search(r'(адрес|сочи|ул\.?|улица|шоссе|пер\.?|переулок|д\.?\s*\d|дом\s*\d)', line, re.I) and not re.search(r'(тел|ключ|фио|квартира|кв\.)', line, re.I):
            address = re.sub(r'^адрес\s*[:\-]?\s*', '', line, flags=re.I).strip()
            break
    if not address:
        m = ADDRESS_RE.search(text or '')
        address = m.group(1).strip(' ,.;') if m else ''
    phones = PHONE_RE.findall(text or '')
    return {'address': address, 'apartment': apt.group(1) if apt else '', 'key_numbers': numbers, 'phones': phones}


def normalize_hex_value(value: str) -> str:
    v = (value or "").strip().upper()
    v = v.replace(" ", "").replace(":", "").replace("-", "")

    if v.startswith("000000") and len(v) == 14:
        v = v[6:]

    return v


def find_key(number_or_hex: str):
    raw = (number_or_hex or "").strip()
    hex_value = normalize_hex_value(raw)

    with db() as conn:
        row = conn.execute(
            "SELECT * FROM keys WHERE number=?",
            (raw,),
        ).fetchone()

        if row:
            return dict(row)

        if re.fullmatch(r"[0-9A-F]{8}", hex_value):
            row = conn.execute(
                "SELECT * FROM keys WHERE UPPER(hex_value)=?",
                (hex_value,),
            ).fetchone()

            if row:
                return dict(row)

            return {
                "number": "",
                "hex_value": hex_value,
                "key_type": "HEX вручную",
            }

    return None


def find_panels_by_address(address: str):
    q = normalize(address)
    with db() as conn:
        rows = conn.execute('SELECT * FROM panels WHERE enabled=1 ORDER BY address, entrance, name').fetchall()
    if not q:
        return []
    result = []
    for r in rows:
        nr = normalize(r['address'])
        if q in nr or nr in q:
            result.append(dict(r))
    return result


def get_panels(panel_ids: Iterable[int] | None = None, tag: str | None = None):
    with db() as conn:
        if panel_ids:
            ids = list(panel_ids)
            ph = ','.join('?' for _ in ids)
            return [dict(r) for r in conn.execute(f'SELECT * FROM panels WHERE enabled=1 AND id IN ({ph}) ORDER BY address,name', ids)]
        if tag:
            return [dict(r) for r in conn.execute('SELECT * FROM panels WHERE enabled=1 AND tags LIKE ? ORDER BY address,name', (f'%{tag}%',))]
        return [dict(r) for r in conn.execute('SELECT * FROM panels WHERE enabled=1 ORDER BY address,name')]


def crm_add_key(mac: str, hex_value: str, flat_num: str, inner: int):
    url = f"{settings.crm_base_url.rstrip('/')}/front/device-keys/{mac}/create-key"
    payload = {'value': hex_value, 'numberSystem': '16', 'flatNum': flat_num or '0', 'inner': int(inner)}
    if settings.dry_run:
        return {'ok': True, 'status': 'DRY_RUN', 'response': json.dumps({'url': url, 'payload': payload}, ensure_ascii=False)}
    if not settings.crm_cookie:
        return {'ok': False, 'status': 'NO_COOKIE', 'response': 'CRM_COOKIE пустой в .env'}
    try:
        r = requests.post(url, headers={'Cookie': settings.crm_cookie, 'Content-Type': 'application/json'}, data=json.dumps(payload), timeout=settings.request_timeout)
        ok = ('true' in r.text.lower()) or r.ok
        return {'ok': ok, 'status': str(r.status_code), 'response': r.text[:1000]}
    except Exception as e:
        return {'ok': False, 'status': 'ERROR', 'response': str(e)}


def write_key_to_panels(
    mode: str,
    key_item: dict,
    panels: list[dict],
    flat_num="0",
    inner=1,
    address="",
):
    results = []

    for p in panels:
        res = crm_add_key(
            p["mac"],
            key_item["hex_value"],
            flat_num,
            inner,
        )

        with db() as conn:
            conn.execute(
                """
                INSERT INTO operation_log(
                    mode,
                    printed_number,
                    hex_value,
                    flat_num,
                    mac,
                    panel_name,
                    status,
                    response,
                    address,
                    apartment
                )
                VALUES(?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    mode,
                    key_item.get("number", ""),
                    key_item["hex_value"],
                    str(flat_num),
                    p["mac"],
                    p.get("name", ""),
                    res["status"],
                    res["response"],
                    address or p.get("address", ""),
                    str(flat_num),
                ),
            )

        results.append({"panel": p, **res})

    return results


def import_keys_file(filename: str, content: bytes) -> int:
    rows = []
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text), delimiter=';' if ';' in text.splitlines()[0] else ',')
        for r in reader:
            rows.append(r)
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        for ws in wb.worksheets:
            headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                rows.append(d)
    added = 0
    with db() as conn:
        for r in rows:
            number = str(r.get('number') or r.get('номер') or r.get('printed_number') or r.get('напечатанный номер') or '').strip()
            hexv = str(r.get('hex') or r.get('hex_value') or r.get('код') or r.get('код для вшития') or '').strip().upper().replace(' ','')
            typ = str(r.get('type') or r.get('тип') or r.get('вид') or '').strip()
            if number and re.fullmatch(r'[0-9A-F]{6,16}', hexv):
                conn.execute('INSERT INTO keys(number,hex_value,key_type,updated_at) VALUES(?,?,?,CURRENT_TIMESTAMP) ON CONFLICT(number) DO UPDATE SET hex_value=excluded.hex_value,key_type=excluded.key_type,updated_at=CURRENT_TIMESTAMP', (number, hexv, typ))
                added += 1
    return added


def import_panels_csv(content: bytes) -> int:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text), delimiter=';' if ';' in text.splitlines()[0] else ',')
    count = 0
    with db() as conn:
        for r in reader:
            address = (r.get('address') or r.get('адрес') or '').strip()
            mac = (r.get('mac') or r.get('MAC') or '').strip().upper()
            name = (r.get('name') or r.get('название') or f'{address} {mac}').strip()
            entrance = (r.get('entrance') or r.get('вход') or r.get('подъезд') or '').strip()
            tags = (r.get('tags') or r.get('теги') or '').strip()
            if address and mac:
                conn.execute('INSERT INTO panels(address,entrance,name,mac,tags) VALUES(?,?,?,?,?) ON CONFLICT(mac) DO UPDATE SET address=excluded.address,entrance=excluded.entrance,name=excluded.name,tags=excluded.tags', (address, entrance, name, mac, tags))
                count += 1
    return count
def universal_search(query: str):
    q = (query or "").strip()
    hex_value = normalize_hex_value(q)

    result = {
        "query": q,
        "key": None,
        "last_operation": None,
        "history": [],
        "address_results": [],
    }

    if not q:
        return result

    key = find_key(q)
    result["key"] = key

    with db() as conn:
        if key:
            history = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT *
                    FROM operation_log
                    WHERE printed_number = ?
                       OR UPPER(hex_value) = ?
                    ORDER BY id DESC
                    LIMIT 50
                    """,
                    (
                        key.get("number", ""),
                        key.get("hex_value", "").upper(),
                    ),
                )
            ]

            result["history"] = history
            result["last_operation"] = history[0] if history else None

        # Поиск по адресу / квартире / номеру ключа / HEX
        result["address_results"] = [
            dict(r)
            for r in conn.execute(
                """
                SELECT *
                FROM operation_log
                WHERE address LIKE ?
                   OR apartment LIKE ?
                   OR flat_num LIKE ?
                   OR printed_number LIKE ?
                   OR UPPER(hex_value) LIKE ?
                ORDER BY id DESC
                LIMIT 50
                """,
                (
                    f"%{q}%",
                    f"%{q}%",
                    f"%{q}%",
                    f"%{q}%",
                    f"%{hex_value}%",
                ),
            )
        ]

    return result
def split_panel_address(full_address: str) -> tuple[str, str]:
    text = (full_address or "").strip()
    text = re.sub(r"\s+", " ", text)

    if "," not in text:
        return text, ""

    parts = [p.strip() for p in text.split(",") if p.strip()]

    address = parts[0]
    entrance = ", ".join(parts[1:])

    return address, entrance


def import_panels_excel(filename: str, content: bytes) -> dict:
    result = {
        "added": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
    }

    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        result["errors"] += 1
        return result

    wb = load_workbook(io.BytesIO(content), data_only=True)

    rows = []

    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h or "").strip().lower() for h in header_row]

        address_idx = None
        mac_idx = None

        for i, h in enumerate(headers):
            if h in ("адрес", "address"):
                address_idx = i
            if h in ("mac", "мас", "мак"):
                mac_idx = i

        if address_idx is None or mac_idx is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_address = str(row[address_idx] or "").strip()
            raw_mac = str(row[mac_idx] or "").strip().upper()

            mac = raw_mac.replace(" ", "")

            if not raw_address or not mac:
                result["skipped"] += 1
                continue

            if not re.fullmatch(r"[0-9A-F]{2}(:[0-9A-F]{2}){5}", mac):
                result["errors"] += 1
                continue

            address, entrance = split_panel_address(raw_address)
            name = f"{address} {entrance}".strip()

            rows.append(
                {
                    "address": address,
                    "entrance": entrance,
                    "name": name,
                    "mac": mac,
                    "tags": "",
                }
            )

    with db() as conn:
        for item in rows:
            existing = conn.execute(
                "SELECT * FROM panels WHERE mac = ?",
                (item["mac"],),
            ).fetchone()

            if existing:
                old = dict(existing)

                changed = (
                    old.get("address") != item["address"]
                    or old.get("entrance") != item["entrance"]
                    or old.get("name") != item["name"]
                )

                conn.execute(
                    """
                    UPDATE panels
                    SET address = ?,
                        entrance = ?,
                        name = ?,
                        tags = ?
                    WHERE mac = ?
                    """,
                    (
                        item["address"],
                        item["entrance"],
                        item["name"],
                        item["tags"],
                        item["mac"],
                    ),
                )

                if changed:
                    result["updated"] += 1
                else:
                    result["skipped"] += 1
            else:
                conn.execute(
                    """
                    INSERT INTO panels(address, entrance, name, mac, tags)
                    VALUES(?,?,?,?,?)
                    """,
                    (
                        item["address"],
                        item["entrance"],
                        item["name"],
                        item["mac"],
                        item["tags"],
                    ),
                )

                result["added"] += 1

    return result