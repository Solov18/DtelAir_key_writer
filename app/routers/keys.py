import io
import logging
import re
from urllib.parse import urlencode

from fastapi import APIRouter, File, Form, Query, Request, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook

from app.repositories.key_repository import (
    ASSIGNMENT_TYPE_NAMES,
    KEY_STATUSES,
    create_key_type,
    get_all_keys_for_export,
    get_key,
    get_key_assignments,
    get_key_history,
    get_key_statistics,
    get_key_type,
    get_key_types,
    get_keys_page,
    get_missing_key_numbers,
    key_status_name,
    prepare_key_range,
    save_key_hex,
    save_prepared_key,
    set_key_assignment,
    set_key_status,
    update_key,
    update_key_assignment,
    update_active_assignment_metadata,
    update_key_type,
)
from app.repositories.log_repository import normalize_operation_row
from app.repositories import key_access_repository
from app.repositories.panel_repository import (
    get_panels_by_ids,
    get_panels_for_exact_address,
    resolve_exact_panel_address,
)
from app.services import import_keys_file
from app.services.audit import log_event
from app.services.excel_export import excel_safe_value, excel_text_cell
from app.services.key_lifecycle import release_key as release_key_lifecycle
from app.services.key_lifecycle import reassign_key as reassign_key_lifecycle
from app.services.key_lifecycle import replace_key as replace_key_lifecycle
from app.repositories.key_lifecycle_repository import get_key_snapshot
from app.services.keys import find_keys
from app.services.writer import write_key_to_panels
from app.templates_config import templates

router = APIRouter()
logger = logging.getLogger("uvicorn.error")


def _user_name(request: Request) -> str:
    user = request.session.get("user", {})
    return user.get("full_name") or user.get("login") or "Система"


def _keys_redirect(**params) -> RedirectResponse:
    clean_params = {
        key: value
        for key, value in params.items()
        if value not in (None, "")
    }
    suffix = f"?{urlencode(clean_params)}" if clean_params else ""
    return RedirectResponse(f"/keys{suffix}", status_code=303)


def _import_report_from_query(request: Request) -> dict | None:
    if request.query_params.get("imported") != "1":
        return None

    names = ("created_types", "added", "updated", "duplicates", "errors")
    return {
        name: int(request.query_params.get(name, "0") or 0)
        for name in names
    }


@router.get("/keys", response_class=HTMLResponse)
def keys_page(
    request: Request,
    q: str = "",
    key_type_id: int = 0,
    status: str = "",
    availability: str = "",
    added_from: str = "",
    added_to: str = "",
    assigned_from: str = "",
    assigned_to: str = "",
    page: int = 1,
    selected_key_id: int = 0,
):
    key_page = get_keys_page(
        query=q,
        key_type_id=key_type_id or None,
        status=status,
        availability=availability,
        added_from=added_from,
        added_to=added_to,
        assigned_from=assigned_from,
        assigned_to=assigned_to,
        page=page,
        page_size=20,
    )

    selected_key = get_key(selected_key_id) if selected_key_id else None
    if not selected_key and key_page["items"]:
        selected_key = get_key(key_page["items"][0]["id"])

    selected_history = []
    if selected_key:
        selected_history = [
            normalize_operation_row(item)
            for item in get_key_history(selected_key["id"], limit=4)
        ]

    filters = {
        "q": q,
        "key_type_id": key_type_id,
        "status": status,
        "availability": availability,
        "added_from": added_from,
        "added_to": added_to,
        "assigned_from": assigned_from,
        "assigned_to": assigned_to,
    }
    base_query = urlencode(
        {
            name: value
            for name, value in filters.items()
            if value not in (None, "", 0)
        }
    )
    row_query = urlencode(
        {
            **{
                name: value
                for name, value in filters.items()
                if value not in (None, "", 0)
            },
            "page": key_page["page"],
        }
    )

    return templates.TemplateResponse(
        "keys.html",
        {
            "request": request,
            "keys": key_page["items"],
            "key_page": key_page,
            "key_types": get_key_types(),
            "active_key_types": get_key_types(include_archived=False),
            "key_statuses": KEY_STATUSES,
            "statistics": get_key_statistics(),
            "filters": filters,
            "selected_key": selected_key,
            "selected_history": selected_history,
            "base_query": base_query,
            "row_query": row_query,
            "row_offset": (key_page["page"] - 1) * key_page["page_size"],
            "import_report": _import_report_from_query(request),
            "message": request.query_params.get("message", ""),
            "error": request.query_params.get("error", ""),
        },
    )


@router.post("/keys/types")
def key_type_create(
    request: Request,
    name: str = Form(...),
    color: str = Form("#2A9DF4"),
    note: str = Form(""),
):
    try:
        key_type_id = create_key_type(name, color, note)
    except ValueError as error:
        return _keys_redirect(error=str(error))

    log_event(
        request=request,
        action="key_type_create",
        object_type="Тип ключа",
        object_name=name.strip(),
        details="Создан новый тип ключа",
    )
    return _keys_redirect(message="Тип ключа создан", edit_type=key_type_id)


@router.post("/keys/types/{key_type_id}")
def key_type_update(
    request: Request,
    key_type_id: int,
    name: str = Form(...),
    color: str = Form("#2A9DF4"),
    note: str = Form(""),
    enabled: str = Form("0"),
):
    try:
        update_key_type(
            key_type_id,
            name,
            color,
            note,
            enabled == "1",
        )
    except ValueError as error:
        return _keys_redirect(error=str(error))

    log_event(
        request=request,
        action="key_type_update",
        object_type="Тип ключа",
        object_name=name.strip(),
        details="Обновлены параметры типа ключа",
    )
    return _keys_redirect(message="Тип ключа обновлён")


@router.post("/keys/prepare", response_class=HTMLResponse)
def keys_prepare(
    request: Request,
    key_type_id: int = Form(...),
    start_number: str = Form(...),
    count: int = Form(...),
):
    try:
        batch = prepare_key_range(
            key_type_id,
            start_number,
            count,
            _user_name(request),
        )
    except ValueError as error:
        return _keys_redirect(error=str(error))

    log_event(
        request=request,
        action="keys_prepare",
        object_type="Партия ключей",
        object_name=f"{batch['key_type']['name']} №{batch['start']}–{batch['end']}",
        details=(
            f"Подготовлено к считыванию: {len(batch['rows'])}; "
            f"уже были готовы: {batch['filled_existing']}. "
            "Записи без HEX не создавались."
        ),
    )

    return templates.TemplateResponse(
        "keys_prepare.html",
        {
            "request": request,
            "batch": batch,
        },
    )


@router.get("/keys/arbitrary", response_class=HTMLResponse)
def keys_arbitrary(
    request: Request,
    key_type_id: int = Query(...),
    count: int = Query(1),
    suggested_number: str = Query(""),
):
    key_type = get_key_type(key_type_id)
    if not key_type or not key_type.get("enabled"):
        return _keys_redirect(error="Выберите активный тип ключа.")
    if count < 1 or count > 500:
        return _keys_redirect(error="За один раз можно добавить от 1 до 500 произвольных ключей.")

    clean_suggestion = suggested_number.strip()
    if clean_suggestion and not re.fullmatch(r"[0-9]+", clean_suggestion):
        return _keys_redirect(error="Предложенный номер должен состоять только из цифр.")

    return templates.TemplateResponse(
        "keys_arbitrary.html",
        {
            "request": request,
            "key_type": key_type,
            "rows": [
                {
                    "index": index + 1,
                    "number": clean_suggestion if index == 0 else "",
                }
                for index in range(count)
            ],
        },
    )


@router.get("/keys/missing", response_class=HTMLResponse)
def keys_missing(
    request: Request,
    key_type_id: int = Query(...),
    start_number: str = Query(""),
    end_number: str = Query(""),
):
    try:
        result = get_missing_key_numbers(
            key_type_id,
            start_number,
            end_number,
        )
    except ValueError as error:
        return _keys_redirect(error=str(error))

    return templates.TemplateResponse(
        "keys_missing.html",
        {
            "request": request,
            "result": result,
            "active_key_types": get_key_types(include_archived=False),
        },
    )


@router.post("/keys/scan")
async def prepared_key_hex_save(request: Request):
    try:
        payload = await request.json()
        allow_replace = bool(payload.get("replace", False))
        key = await run_in_threadpool(
            save_prepared_key,
            int(payload.get("key_type_id", 0)),
            str(payload.get("number", "")),
            str(payload.get("hex_value", "")),
            _user_name(request),
            allow_replace=allow_replace,
        )
    except (ValueError, TypeError) as error:
        return JSONResponse(
            {"ok": False, "error": str(error)},
            status_code=400,
        )

    await run_in_threadpool(
        log_event,
        request=request,
        action="key_hex_scan",
        object_type="Ключ",
        object_name=f"{key['type_name']} №{key['number']}",
        details=(
            "HEX исправлен и сохранён"
            if allow_replace
            else "Ключ создан вместе с HEX"
        ),
        key_id=key["id"],
        key_type=key["type_name"],
        printed_number=key["number"],
        hex_value=key["hex_value"],
    )
    return JSONResponse(
        {
            "ok": True,
            "key_id": key["id"],
            "hex_value": key["hex_value"],
            "message": (
                "Исправление сохранено"
                if allow_replace
                else "Ключ и HEX сохранены автоматически"
            ),
        }
    )


@router.post("/keys/{key_id}/hex")
async def key_hex_save(request: Request, key_id: int):
    try:
        payload = await request.json()
        allow_replace = bool(payload.get("replace", False))
        key = await run_in_threadpool(
            save_key_hex,
            key_id,
            str(payload.get("hex_value", "")),
            _user_name(request),
            allow_replace=allow_replace,
        )
    except (ValueError, TypeError) as error:
        return JSONResponse(
            {"ok": False, "error": str(error)},
            status_code=400,
        )

    await run_in_threadpool(
        log_event,
        request=request,
        action="key_hex_scan",
        object_type="Ключ",
        object_name=f"{key['type_name']} №{key['number']}",
        details=(
            "HEX исправлен и сохранён"
            if allow_replace
            else "HEX считан и сохранён"
        ),
        key_id=key_id,
        key_type=key["type_name"],
        printed_number=key["number"],
        hex_value=key["hex_value"],
    )
    return JSONResponse(
        {
            "ok": True,
            "key_id": key_id,
            "hex_value": key["hex_value"],
            "message": "Исправление сохранено" if allow_replace else "Сохранено автоматически",
        }
    )


@router.post("/keys/import")
async def keys_import(request: Request, file: UploadFile = File(...)):
    filename = file.filename or ""
    content = await file.read()
    report = await run_in_threadpool(
        import_keys_file,
        filename,
        content,
        _user_name(request),
    )

    await run_in_threadpool(
        log_event,
        request=request,
        action="import_keys",
        object_type="Файл ключей",
        object_name=file.filename or "Импорт",
        details=(
            f"Типов создано: {report['created_types']}; "
            f"ключей добавлено: {report['added']}; "
            f"обновлено: {report['updated']}; "
            f"дубликатов: {report['duplicates']}; "
            f"ошибок: {report['errors']}"
        ),
        status="success" if report["errors"] == 0 else "warning",
    )

    return _keys_redirect(imported=1, **{
        key: report[key]
        for key in ("created_types", "added", "updated", "duplicates", "errors")
    })


@router.get("/keys/export")
def keys_export():
    # Write-only worksheets keep memory usage bounded when the registry
    # contains tens of thousands of keys.
    workbook = Workbook(write_only=True)
    used_names: set[str] = set()

    keys_by_type: dict[str, list[dict]] = {}
    for key in get_all_keys_for_export():
        keys_by_type.setdefault(key["type_name"], []).append(key)

    for type_name, items in keys_by_type.items():
        base_name = re.sub(r"[\\/*?:\[\]]", "_", type_name)[:31] or "Без типа"
        sheet_name = base_name
        suffix = 2
        while sheet_name.lower() in used_names:
            marker = f"_{suffix}"
            sheet_name = f"{base_name[:31 - len(marker)]}{marker}"
            suffix += 1
        used_names.add(sheet_name.lower())

        sheet = workbook.create_sheet(sheet_name)
        sheet.freeze_panes = "A2"
        sheet.append(
            [
                "Номер",
                "HEX",
                "Статус",
                "Комментарий",
                "Дата добавления",
                "Кем добавлен",
            ]
        )
        for key in items:
            sheet.append(
                [
                    excel_text_cell(sheet, key["number"]),
                    key["hex_value"],
                    key_status_name(key["status"]),
                    key["note"],
                    excel_safe_value(key["created_at"]),
                    key["created_by"],
                ]
            )
        sheet.auto_filter.ref = f"A1:F{len(items) + 1}"

    if not workbook.worksheets:
        sheet = workbook.create_sheet("Ключи")
        sheet.append(
            [
                "Номер",
                "HEX",
                "Статус",
                "Комментарий",
                "Дата добавления",
                "Кем добавлен",
            ]
        )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:F1"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="keys_export.xlsx"'
        },
    )


@router.get("/keys/panels-for-address")
def key_assignment_panels_for_address(address: str = Query("")):
    canonical = resolve_exact_panel_address(address)
    if not canonical:
        return JSONResponse(
            {"ok": False, "error": "Адрес не найден в CRM", "panels": []},
            status_code=404,
        )
    panels = get_panels_for_exact_address(canonical)
    return {
        "ok": True,
        "address": canonical,
        "panels": [
            {
                "id": int(panel["id"]),
                "address": panel.get("address") or "",
                "entrance": panel.get("entrance") or panel.get("name") or "",
                "name": panel.get("name") or "",
                "mac": panel.get("mac") or "",
                "status": panel.get("network_status") or "unknown",
                "status_name": panel.get("status_name") or "Не проверено",
            }
            for panel in panels
        ],
    }


@router.get("/keys/{key_id}", response_class=HTMLResponse)
def key_detail(request: Request, key_id: int):
    key = get_key(key_id)
    if not key:
        return _keys_redirect(error="Ключ не найден")
    history = [
        normalize_operation_row(item)
        for item in get_key_history(key_id)
    ]

    return templates.TemplateResponse(
        "key_detail.html",
        {
            "request": request,
            "key": key,
            "key_types": get_key_types(),
            "key_statuses": KEY_STATUSES,
            "assignments": get_key_assignments(key_id),
            "active_accesses": key_access_repository.get_active_accesses_with_panels(key_id),
            "assignment_type_names": ASSIGNMENT_TYPE_NAMES,
            "history": history,
            "assignment_changes": [
                item
                for item in history
                if item.get("action") == "key_assignment_update"
            ],
            "message": request.query_params.get("message", ""),
            "error": request.query_params.get("error", ""),
        },
    )


@router.post("/keys/{key_id}/update")
def key_update_route(
    request: Request,
    key_id: int,
    key_type_id: int = Form(...),
    number: str = Form(...),
    hex_value: str = Form(""),
    note: str = Form(""),
):
    try:
        update_key(
            key_id,
            key_type_id,
            number,
            hex_value,
            note,
        )
    except ValueError as error:
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': str(error)})}",
            status_code=303,
        )

    key = get_key(key_id)
    log_event(
        request=request,
        action="key_update",
        object_type="Ключ",
        object_name=f"{(key or {}).get('type_name', '')} №{number}",
        details="Карточка ключа обновлена",
        key_id=key_id,
        key_type=(key or {}).get("type_name", ""),
        printed_number=number,
        hex_value=(key or {}).get("hex_value", "-"),
    )
    return RedirectResponse(
        f"/keys/{key_id}?message=Ключ+обновлён",
        status_code=303,
    )


@router.post("/keys/{key_id}/assignment")
def key_assignment_update_route(
    request: Request,
    key_id: int,
    assignment_type: str = Form(...),
    address: str = Form(...),
    apartment: str = Form(""),
    owner_name: str = Form(""),
    reason: str = Form(...),
    target_panel_ids: list[int] = Form([]),
    assignment_action: str = Form("reassign"),
):
    # FastAPI supplies a list for real form requests.  Keep direct route calls
    # in tests and internal tools safe from the Form() descriptor default.
    target_panel_ids = (
        [int(value) for value in target_panel_ids]
        if isinstance(target_panel_ids, (list, tuple))
        else []
    )
    assignment_action = (
        assignment_action if isinstance(assignment_action, str) else "reassign"
    )
    snapshot = get_key_snapshot(key_id)
    partial_warning = ""
    try:
        canonical_address = resolve_exact_panel_address(address)
        if not canonical_address:
            raise ValueError("Адрес не найден в CRM. Выберите существующий адрес.")
        address = canonical_address
        if assignment_action not in {"add_access", "reassign"}:
            raise ValueError("Выберите действие: добавить доступ или переназначить ключ.")
        active_accesses = list(snapshot.get("accesses") or []) if snapshot else []
        old_assignment = dict((snapshot.get("assignments") or [{}])[0]) if snapshot else {}
        primary_access = next(
            (dict(item) for item in active_accesses if item.get("is_primary")),
            dict(active_accesses[0]) if active_accesses else old_assignment,
        )
        physical_change = bool(primary_access) and (
            address != str(primary_access.get("address") or "")
            or str(apartment or "") != str(primary_access.get("apartment") or "")
        )
        requires_panel_write = assignment_action == "add_access" or physical_change
        if requires_panel_write:
            if not target_panel_ids:
                raise ValueError("Выберите хотя бы одну панель нового адреса.")
            target_panels = get_panels_by_ids(target_panel_ids)
            if len(target_panels) != len(set(target_panel_ids)):
                raise ValueError("Одна или несколько выбранных панелей не найдены.")
            allowed_panel_ids = {
                int(panel["id"])
                for panel in get_panels_for_exact_address(address)
            }
            foreign_panel_ids = set(target_panel_ids) - allowed_panel_ids
            if foreign_panel_ids:
                raise ValueError(
                    "Выбранные панели не относятся к указанному адресу. "
                    "Обновите список панелей после выбора адреса."
                )
            if any(not panel.get("enabled") or not (panel.get("mac") or "").strip() for panel in target_panels):
                raise ValueError("Среди выбранных есть отключённая панель или панель без MAC.")
            key_before = get_key(key_id)
            previous_description = "; ".join(
                ", ".join(filter(None, [
                    str(item.get("address") or ""),
                    f"кв. {item.get('apartment')}" if item.get("apartment") else "",
                ]))
                for item in active_accesses
            ) or str(primary_access.get("address") or "")

            def write_target(policy: str, panel_ids: list[int] | None = None):
                selected_ids = set(panel_ids or target_panel_ids)
                selected_panels = [
                    panel for panel in target_panels
                    if int(panel["id"]) in selected_ids
                ]
                return write_key_to_panels(
                    "assignment_update", key_before, selected_panels,
                    flat_num=apartment, inner=1, address=address,
                    request=request, assignment_type=assignment_type,
                    assignment_policy=policy,
                    known_panel_ids=set(),
                    write_option=("add_selected_panels" if policy == "preserve" else "reassign_to_new_address"),
                    previous_assignment=previous_description,
                )

            if assignment_action == "add_access":
                write_results = write_target("preserve")
                lifecycle_result = {"status": "SUCCESS", "write": write_results}
            else:
                lifecycle_result = reassign_key_lifecycle(
                    key_id,
                    write_callback=lambda _old_snapshot, pending_ids: write_target(
                        "replace", pending_ids
                    ),
                    reason=reason, request=request,
                    target_context={
                        "address": address,
                        "apartment": str(apartment or ""),
                        "panel_ids": sorted(set(target_panel_ids)),
                    },
                )
                if lifecycle_result.get("write") is None:
                    failures = [
                        item for item in lifecycle_result.get("release", {}).get("results", [])
                        if not item.get("success")
                    ]
                    detail = "; ".join(
                        filter(None, [
                            f"{(item.get('panel') or {}).get('address', '')} {(item.get('panel') or {}).get('entrance', '')}: "
                            f"{item.get('response') or item.get('status') or 'ошибка'}"
                            for item in failures[:3]
                        ])
                    )
                    raise ValueError(
                        "Назначение не изменено: центральная CRM не подтвердила удаление прежней записи ключа."
                        + (f" {detail}" if detail else "")
                    )
            if not any(
                item.get("written") or item.get("status") in {"SUCCESS", "ALREADY_EXISTS", "ALREADY_ON_PANEL"}
                for item in lifecycle_result.get("write") or []
            ):
                raise ValueError(
                    "Центральная CRM удалила прежнюю запись, но не подтвердила создание записи "
                    "с новым адресом и квартирой. Повторите операцию."
                )
            if lifecycle_result.get("status") == "PARTIAL":
                failed_count = sum(
                    1
                    for item in lifecycle_result.get("write") or []
                    if not (
                        item.get("written")
                        or item.get("status") in {"SUCCESS", "ALREADY_EXISTS", "ALREADY_ON_PANEL"}
                    )
                )
                partial_warning = (
                    "Назначение изменено, но центральная CRM подтвердила не все новые точки доступа. "
                    f"Не завершено: {failed_count}. Повторите операцию для них."
                )
            if assignment_action == "reassign":
                update_active_assignment_metadata(
                    key_id,
                    owner_name=owner_name,
                    assigned_by=_user_name(request),
                )

        if requires_panel_write:
            old = dict(primary_access)
            old_assignment_type = str(
                old.get("assignment_type") or old.get("access_type") or ""
            )
            old_description = ", ".join(filter(None, [
                ASSIGNMENT_TYPE_NAMES.get(old_assignment_type, old_assignment_type),
                old.get("address", ""),
                f"квартира {old.get('apartment')}" if old.get("apartment") else "квартира не указана",
            ]))
            new = {
                "assignment_type": assignment_type, "address": address,
                "apartment": apartment, "owner_name": owner_name,
            }
            new_description = ", ".join(filter(None, [
                ASSIGNMENT_TYPE_NAMES.get(assignment_type, assignment_type), address,
                f"квартира {apartment}" if apartment else "квартира не указана", owner_name,
            ]))
            change = {
                "old": old, "new": new, "old_description": old_description,
                "new_description": new_description, "reason": reason,
            }
            if assignment_action == "add_access":
                # The writer has persisted the additional key_access and panel
                # relations.  Keep the primary assignment untouched.
                access_id = key_access_repository.ensure_access(
                    key_id, assignment_type=assignment_type, address=address,
                    apartment=apartment, owner_name=owner_name, primary=False,
                    source="add_access", created_by=_user_name(request), note=reason,
                )
                key_access_repository.attach_panels(key_id, access_id, target_panel_ids)
        else:
            change = update_key_assignment(
                key_id, assignment_type, address=address, apartment=apartment,
                owner_name=owner_name, reason=reason, assigned_by=_user_name(request),
            )
            key_access_repository.sync_primary_access(
                key_id,
                assignment_type=assignment_type,
                address=address,
                apartment=apartment,
                owner_name=owner_name,
                assignment_id=(get_key(key_id) or {}).get("assignment_id"),
                created_by=_user_name(request),
                note=reason,
            )
    except ValueError as error:
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': str(error)})}#assignment-edit",
            status_code=303,
        )
    except Exception:
        logger.exception("key_assignment.update.failed key_id=%s", key_id)
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': 'Не удалось изменить назначение. Повторите операцию или обратитесь к администратору.'})}#assignment-edit",
            status_code=303,
        )

    key = get_key(key_id)
    details = (
        (
            "Добавлен дополнительный доступ без изменения текущего назначения: "
            f"{change['new_description']}. "
        )
        if assignment_action == "add_access"
        else (
            "Назначение изменено: "
            f"{change['old_description']} → {change['new_description']}. "
        )
    ) + f"Причина: {change['reason']}"
    log_event(
        request=request,
        action="key_assignment_update",
        object_type="Ключ",
        object_name=f"{(key or {}).get('type_name', '')} №{(key or {}).get('number', key_id)}",
        details=details,
        key_id=key_id,
        key_type=(key or {}).get("type_name", ""),
        printed_number=(key or {}).get("number", ""),
        hex_value=(key or {}).get("hex_value", "-"),
        address=change["new"].get("address", ""),
        apartment=change["new"].get("apartment", ""),
        comment=change["reason"],
    )
    response_params = (
        {"error": partial_warning}
        if partial_warning
        else {
            "message": (
                "Дополнительный доступ добавлен"
                if assignment_action == "add_access"
                else "Назначение изменено"
            )
        }
    )
    return RedirectResponse(
        f"/keys/{key_id}?{urlencode(response_params)}",
        status_code=303,
    )


@router.post("/keys/{key_id}/replace")
def key_replace_route(
    request: Request,
    key_id: int,
    new_key_query: str = Form(...),
    new_key_type_id: int = Form(0),
    final_old_status: str = Form("replaced"),
    reason: str = Form(...),
):
    """Replace one physical key while preserving its complete access set."""
    try:
        matches = find_keys(new_key_query, new_key_type_id or None)
        if not matches:
            raise ValueError("Новый ключ не найден в базе CRM.")
        if len(matches) != 1:
            raise ValueError("Номер встречается в нескольких типах. Выберите тип нового ключа.")
        new_key = get_key(int(matches[0]["id"]))
        if not new_key or int(new_key["id"]) == int(key_id):
            raise ValueError("Для замены выберите другой физический ключ.")
        if not (new_key.get("hex_value") or "").strip():
            raise ValueError("У нового ключа отсутствует HEX.")
        new_snapshot = get_key_snapshot(int(new_key["id"]))
        if not new_snapshot or new_snapshot.get("occupied"):
            raise ValueError("Новый ключ уже используется. Для замены нужен свободный ключ.")
        if final_old_status not in {"replaced", "lost", "defective", "archived"}:
            raise ValueError("Выберите корректное состояние старого ключа.")

        def write_replacement(replacement: dict, snapshot: dict) -> list[dict]:
            results: list[dict] = []
            grouped: dict[tuple[str, str, str], list[dict]] = {}
            for item in snapshot.get("panels") or []:
                group_key = (
                    str(item.get("access_address") or item.get("address") or ""),
                    str(item.get("access_apartment") or item.get("flat_num") or ""),
                    str(item.get("access_type") or "resident"),
                )
                grouped.setdefault(group_key, []).append(item)
            for (address, apartment, access_type), panel_rows in grouped.items():
                panel_ids = [int(item["panel_id"]) for item in panel_rows]
                panels = get_panels_by_ids(panel_ids)
                results.extend(write_key_to_panels(
                    "key_replacement", replacement, panels,
                    flat_num=apartment, inner=1, address=address,
                    request=request, assignment_type="",
                    assignment_policy="preserve", known_panel_ids=set(),
                    write_option="replace_physical_key",
                    previous_assignment=f"Замена ключа №{(get_key(key_id) or {}).get('number', key_id)}",
                ))
            return results

        result = replace_key_lifecycle(
            key_id, new_key,
            final_old_status=final_old_status,
            write_callback=write_replacement,
            reason=reason,
            request=request,
        )
        if not result.get("ok"):
            raise ValueError(
                "Замена выполнена не полностью. Состояние операции сохранено; "
                "повторный запуск продолжит только незавершённые шаги."
            )
    except ValueError as error:
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': str(error)})}#key-replace",
            status_code=303,
        )
    except Exception:
        logger.exception("key.replace.failed old_key_id=%s", key_id)
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': 'Не удалось заменить физический ключ.'})}#key-replace",
            status_code=303,
        )

    log_event(
        request=request,
        action="key_replace",
        object_type="Ключ",
        object_name=f"№{(get_key(key_id) or {}).get('number', key_id)} → №{new_key.get('number', '')}",
        details=f"Физический ключ заменён. Причина: {reason}",
        key_id=int(new_key["id"]),
        key_type=new_key.get("type_name", ""),
        printed_number=new_key.get("number", ""),
        hex_value=new_key.get("hex_value", "-"),
        comment=reason,
    )
    return RedirectResponse(
        f"/keys/{int(new_key['id'])}?{urlencode({'message': 'Физический ключ заменён; активные доступы перенесены'})}",
        status_code=303,
    )


@router.post("/keys/{key_id}/status")
def key_status_route(
    request: Request,
    key_id: int,
    status: str = Form(...),
    note: str = Form(""),
):
    try:
        if status in {"free", "lost", "defective", "blocked", "archived"}:
            result = release_key_lifecycle(
                key_id,
                reason=note or f"Статус изменён на {key_status_name(status)}",
                final_status=status,
                request=request,
            )
            if not result["ok"]:
                failed = sum(1 for item in result["results"] if not item["success"])
                raise ValueError(f"Не удалось удалить {failed} записей ключа из CRM. Локальный статус не изменён.")
        else:
            set_key_status(key_id, status, note)
    except ValueError as error:
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': str(error)})}",
            status_code=303,
        )

    key = get_key(key_id)
    log_event(
        request=request,
        action="key_status_change",
        object_type="Ключ",
        object_name=(key or {}).get("number") or str(key_id),
        details=f"Новый статус: {key_status_name(status)}",
        key_id=key_id,
        key_type=(key or {}).get("type_name", ""),
        printed_number=(key or {}).get("number", ""),
        hex_value=(key or {}).get("hex_value", "-"),
        comment=note,
    )
    return RedirectResponse(
        f"/keys/{key_id}?message=Статус+обновлён",
        status_code=303,
    )


@router.post("/keys/{key_id}/release")
def key_release_route(
    request: Request,
    key_id: int,
    note: str = Form("Освобождён вручную"),
):
    try:
        result = release_key_lifecycle(key_id, reason=note, request=request)
        if not result["ok"]:
            failed = sum(1 for item in result["results"] if not item["success"])
            raise ValueError(
                f"Освобождение не завершено: центральная CRM не подтвердила удаление "
                f"{failed} записей ключа. Повторите операцию."
            )
    except ValueError as error:
        return RedirectResponse(
            f"/keys/{key_id}?{urlencode({'error': str(error)})}",
            status_code=303,
        )

    key = get_key(key_id)
    log_event(
        request=request,
        action="key_release",
        object_type="Ключ",
        object_name=(key or {}).get("number") or str(key_id),
        details=note,
        key_id=key_id,
        key_type=(key or {}).get("type_name", ""),
        printed_number=(key or {}).get("number", ""),
        hex_value=(key or {}).get("hex_value", "-"),
    )
    return RedirectResponse(
        f"/keys/{key_id}?message=Ключ+освобождён",
        status_code=303,
    )
