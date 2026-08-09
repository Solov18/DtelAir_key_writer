import asyncio
import io
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import Mock, patch

import requests
from openpyxl import load_workbook
from starlette.requests import Request

from tests.postgres_test_case import PostgreSQLTestCase

from app.repositories import panel_repository
from app.routers import panels as panels_router
from app.services.importer import import_panels_excel
from app.services import panel_api


class PanelRepositoryTests(PostgreSQLTestCase):
    def setUp(self):
        super().setUp()

    def _create(self, address, entrance, mac, ip=""):
        panel_repository.create_or_update_panel(
            address=address,
            entrance=entrance,
            mac=mac,
            ip=ip,
        )
        normalized_mac = panel_repository.normalize_mac(mac)
        return next(
            panel
            for panel in panel_repository.get_all_panels()
            if panel["mac"] == normalized_mac
        )

    def test_panel_export_removes_timezone_after_localizing_datetimes(self):
        panel = {
            "id": 15,
            "address": "ул. Тепличная, д.83",
            "entrance": "Подъезд 4",
            "name": "Панель 4",
            "ip": "10.0.0.15",
            "mac": "08:13:CD:00:17:1E",
            "status_name": "В сети",
            "supply_voltage": 12.9,
            "device_model": "Sokol",
            "firmware_version": "2.2.5.10.5",
            "last_checked_at": datetime(2026, 8, 1, 11, 40, tzinfo=timezone.utc),
            "last_online_at": datetime(2026, 8, 1, 11, 39, tzinfo=timezone.utc),
        }

        with patch.object(panels_router, "get_all_panels", return_value=[panel]):
            response = panels_router.panels_export()

        async def read_response() -> bytes:
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(chunk)
            return b"".join(chunks)

        workbook = load_workbook(io.BytesIO(asyncio.run(read_response())))
        sheet = workbook.active
        checked_at = sheet.cell(row=2, column=11).value
        online_at = sheet.cell(row=2, column=12).value

        self.assertEqual(checked_at, datetime(2026, 8, 1, 14, 40))
        self.assertEqual(online_at, datetime(2026, 8, 1, 14, 39))
        self.assertIsNone(checked_at.tzinfo)
        self.assertIsNone(online_at.tzinfo)

    def test_exported_workbook_can_be_edited_and_imported_without_data_loss(self):
        existing = self._create(
            "ул. Тепличная, д.83",
            "Подъезд 4",
            "08:13:CD:00:17:1E",
            "10.90.171.20",
        )

        response = panels_router.panels_export()

        async def read_response() -> bytes:
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(chunk)
            return b"".join(chunks)

        workbook = load_workbook(io.BytesIO(asyncio.run(read_response())))
        sheet = workbook.active
        sheet.append(
            [
                None,
                "ул. Малышева, д.7",
                "Подъезд 2",
                "ул. Малышева, д.7 Подъезд 2",
                "10.80.72.10",
                "08:13:CD:00:34:81",
            ]
        )
        content = io.BytesIO()
        workbook.save(content)

        report = import_panels_excel("panels_export.xlsx", content.getvalue())

        self.assertEqual(report["added"], 1)
        self.assertEqual(report["errors"], 0)
        panels = {panel["mac"]: panel for panel in panel_repository.get_all_panels()}
        existing_after = panels[existing["mac"]]
        added = panels["08:13:CD:00:34:81"]
        self.assertEqual(existing_after["entrance"], "Подъезд 4")
        self.assertEqual(existing_after["ip"], "10.90.171.20")
        self.assertEqual(added["entrance"], "Подъезд 2")
        self.assertEqual(added["ip"], "10.80.72.10")

    def test_server_filters_pagination_and_status_statistics(self):
        first = self._create("Тепличная 63", "Подъезд 1", "08:13:CD:00:00:01", "10.0.0.1")
        second = self._create("Тепличная 63", "Подъезд 2", "08:13:CD:00:00:02", "10.0.0.2")
        third = self._create("Горького 45", "Главный вход", "08:13:CD:00:00:03")

        panel_repository.update_panel_api_status(
            first["id"],
            {
                "status": "online",
                "response_time_ms": 42,
                "device_model": "ISCom X1",
                "firmware_version": "2.5.0.14.7",
                "temperature": 56.5,
                "supply_voltage": 12.21,
                "uptime_seconds": 90061,
                "sip_registered": True,
                "reported_mac": first["mac"],
                "last_error": "",
            },
        )
        panel_repository.update_panel_api_status(
            second["id"],
            {"status": "offline", "last_error": "Нет соединения"},
        )
        panel_repository.set_panel_enabled(third["id"], False)

        page = panel_repository.get_panel_page(
            query="10.0.0.1",
            status="online",
            address="Тепличная 63",
            page=1,
            page_size=20,
        )
        self.assertEqual(page["total"], 1)
        self.assertEqual(page["items"][0]["id"], first["id"])
        self.assertEqual(page["items"][0]["uptime_text"], "1 дн. 01:01")
        self.assertEqual(page["items"][0]["supply_voltage"], 12.21)
        self.assertTrue(page["items"][0]["mac_matches"])

        punctuation_page = panel_repository.get_panel_page(
            query="08.13-cd 00:00:01",
        )
        self.assertEqual(punctuation_page["total"], 1)
        self.assertEqual(punctuation_page["items"][0]["id"], first["id"])

        combined_page = panel_repository.get_panel_page(
            query="Тепличная 63 подъезд 2",
        )
        self.assertEqual(combined_page["total"], 1)
        self.assertEqual(combined_page["items"][0]["id"], second["id"])

        statistics = panel_repository.get_panel_statistics()
        self.assertEqual(statistics["total"], 3)
        self.assertEqual(statistics["online"], 1)
        self.assertEqual(statistics["offline"], 1)
        self.assertEqual(statistics["disabled"], 1)
        self.assertEqual(statistics["unchecked"], 0)

    def test_sip_problem_filter_returns_only_failed_registered_panels(self):
        failed = self._create("Абрикосовая 21", "Подъезд 1", "08:13:CD:00:40:01", "10.0.40.1")
        healthy = self._create("Абрикосовая 21", "Подъезд 2", "08:13:CD:00:40:02", "10.0.40.2")
        unknown = self._create("Абрикосовая 21", "Подъезд 3", "08:13:CD:00:40:03", "10.0.40.3")

        panel_repository.update_panel_api_status(failed["id"], {"status": "online", "sip_registered": False})
        panel_repository.update_panel_api_status(healthy["id"], {"status": "online", "sip_registered": True})
        panel_repository.update_panel_api_status(unknown["id"], {"status": "online"})

        page = panel_repository.get_panel_page(status="sip_error")

        self.assertEqual(page["total"], 1)
        self.assertEqual(page["items"][0]["id"], failed["id"])
        self.assertEqual(panel_repository.get_panel_statistics()["sip_failed"], 1)

    def test_voltage_problem_filter_uses_monitoring_thresholds(self):
        low = self._create("Волжская 38", "Калитка", "08:13:CD:00:41:01", "10.0.41.1")
        normal = self._create("Волжская 38", "Подъезд 1", "08:13:CD:00:41:02", "10.0.41.2")
        panel_repository.update_panel_api_status(low["id"], {"status": "online", "supply_voltage": 9.34})
        panel_repository.update_panel_api_status(normal["id"], {"status": "online", "supply_voltage": 12.9})

        page = panel_repository.get_panel_page(status="voltage_alert")

        self.assertEqual(page["total"], 1)
        self.assertEqual(page["items"][0]["id"], low["id"])

    def test_unchecked_and_voltage_boundaries_are_not_reported_as_offline(self):
        unchecked = self._create("Мира 8", "1", "08:13:CD:00:00:07")
        item = panel_repository.get_panel_by_id(unchecked["id"])
        self.assertEqual(item["network_status"], "unknown")
        self.assertEqual(item["status_name"], "Не проверено")
        self.assertEqual(panel_repository.get_panel_statistics()["offline"], 0)

        for value, expected in (
            (None, "missing"),
            (12.79, "alert"),
            (12.8, "normal"),
            (13.5, "normal"),
            (13.51, "alert"),
        ):
            row = {**unchecked, "supply_voltage": value}
            normalized = panel_repository.normalize_panel_row(row)
            self.assertEqual(normalized["voltage_tone"], expected)

    def test_page_reads_cached_database_values_without_panel_api_calls(self):
        item = self._create("Курортный 75Д", "Калитка", "08:13:CD:00:00:08")
        scope = {
            "type": "http",
            "http_version": "1.1",
            "method": "GET",
            "scheme": "http",
            "path": "/panels",
            "raw_path": b"/panels",
            "query_string": b"",
            "headers": [],
            "client": ("127.0.0.1", 1234),
            "server": ("test", 80),
            "session": {
                "user": {
                    "id": 1,
                    "login": "admin",
                    "full_name": "Администратор",
                    "permissions": ["view", "manage_panels"],
                }
            },
        }
        request = Request(scope)
        rendered = Mock()
        with (
            patch.object(panels_router.templates, "TemplateResponse", return_value=rendered) as template,
            patch.object(panels_router, "_is_admin", return_value=True),
            patch("app.routers.panels.check_panel") as checker,
        ):
            result = panels_router.panels_page(request)
        self.assertIs(result, rendered)
        checker.assert_not_called()
        context = template.call_args.args[1]
        self.assertEqual(context["panels"][0]["id"], item["id"])
        self.assertIn("monitor_state", context)

    def test_status_cache_keeps_last_successful_device_data_on_failure(self):
        item = self._create("Лесная 12", "1", "08:13:CD:00:00:04", "10.0.0.4")
        panel_repository.update_panel_api_status(
            item["id"],
            {
                "status": "online",
                "device_model": "ISCom X1",
                "firmware_version": "2.5.0",
                "last_error": "",
            },
        )
        panel_repository.update_panel_api_status(
            item["id"],
            {"status": "offline", "last_error": "Тайм-аут"},
        )

        cached = panel_repository.get_panel_by_id(item["id"])
        self.assertEqual(cached["network_status"], "offline")
        self.assertEqual(cached["device_model"], "ISCom X1")
        self.assertEqual(cached["firmware_version"], "2.5.0")
        self.assertEqual(cached["last_error"], "Тайм-аут")


class PanelApiTests(unittest.TestCase):
    def setUp(self):
        self.panel = {"ip": "10.10.1.15"}

    @staticmethod
    def _response(*, status=200, payload=None, content=None, content_type="application/json"):
        response = Mock()
        response.status_code = status
        if content is None:
            content = b"{}" if payload is not None else b""
        response.content = content
        response.headers = {"Content-Type": content_type}
        response.json.return_value = payload if payload is not None else {}
        return response

    def test_check_uses_common_basic_auth_and_collects_real_fields(self):
        responses = [
            self._response(
                payload={
                    "model": "GK7205V300",
                    "temperature": 56.5,
                    "mac": "08:13:CD:00:00:01",
                    "deviceModel": "ISCom X1 (rev.5)",
                    "uptime": 1001,
                    "registerStatus": True,
                }
            ),
            self._response(payload={"power": {"dc": 12.21}, "chipId": 1000000000000000001}),
            self._response(payload={"opt": {"name": "2.5.0.14.7"}}),
        ]
        with (
            patch.object(panel_api.settings, "panel_api_login", "common-user"),
            patch.object(panel_api.settings, "panel_api_password", "common-password"),
            patch.object(panel_api.settings, "panel_api_timeout", 2.5),
            patch("app.services.panel_api._http_request", side_effect=responses) as request_mock,
        ):
            result = panel_api.check_panel(self.panel)

        self.assertEqual(result["status"], "online")
        self.assertEqual(result["device_model"], "ISCom X1 (rev.5)")
        self.assertEqual(result["firmware_version"], "2.5.0.14.7")
        self.assertEqual(result["temperature"], 56.5)
        self.assertEqual(result["supply_voltage"], 12.21)
        self.assertTrue(result["sip_registered"])
        self.assertEqual(request_mock.call_count, 3)
        first_call = request_mock.call_args_list[0]
        self.assertEqual(first_call.args[:2], ("GET", "http://10.10.1.15/system/info"))
        self.assertEqual(first_call.kwargs["auth"].username, "common-user")
        self.assertEqual(first_call.kwargs["auth"].password, "common-password")
        self.assertEqual(first_call.kwargs["timeout"], 2.5)
        self.assertEqual(
            request_mock.call_args_list[1].args[:2],
            ("GET", "http://10.10.1.15/v1/mcu/info"),
        )

    def test_check_maps_timeout_and_bad_credentials_to_clear_statuses(self):
        with (
            patch.object(panel_api.settings, "panel_api_login", "user"),
            patch.object(panel_api.settings, "panel_api_password", "password"),
            patch("app.services.panel_api._http_request", side_effect=requests.Timeout()),
        ):
            timeout_result = panel_api.check_panel(self.panel)
        self.assertEqual(timeout_result["status"], "offline")

        with (
            patch.object(panel_api.settings, "panel_api_login", "user"),
            patch.object(panel_api.settings, "panel_api_password", "password"),
            patch(
                "app.services.panel_api._http_request",
                return_value=self._response(status=401),
            ),
        ):
            auth_result = panel_api.check_panel(self.panel)
        self.assertEqual(auth_result["status"], "auth_error")

    def test_snapshot_and_reboot_use_documented_endpoints(self):
        snapshot = self._response(
            content=b"jpeg-data",
            content_type="image/jpeg",
        )
        reboot = self._response(content=b"", content_type="text/plain")
        with (
            patch.object(panel_api.settings, "panel_api_login", "user"),
            patch.object(panel_api.settings, "panel_api_password", "password"),
            patch(
                "app.services.panel_api._http_request",
                side_effect=[snapshot, reboot],
            ) as request_mock,
        ):
            content, content_type = panel_api.get_panel_snapshot(self.panel)
            panel_api.reboot_panel(self.panel)

        self.assertEqual(content, b"jpeg-data")
        self.assertEqual(content_type, "image/jpeg")
        self.assertEqual(request_mock.call_args_list[0].args[:2], ("GET", "http://10.10.1.15/camera/snapshot"))
        self.assertEqual(request_mock.call_args_list[1].args[:2], ("PUT", "http://10.10.1.15/system/reboot"))

    def test_registry_template_has_monitoring_columns_and_no_eager_camera(self):
        panels_router.templates.env.get_template("panels.html")
        source = Path("app/templates/panels.html").read_text(encoding="utf-8")
        self.assertIn("<th>Температура</th>", source)
        self.assertIn("<th>Прошивка</th>", source)
        self.assertNotIn("<th>Панель</th>", source)
        self.assertNotIn("<th>Последний онлайн</th>", source)
        self.assertNotIn('<img src="/panels/', source)
        self.assertIn("loadPanelSnapshot", source)


if __name__ == "__main__":
    unittest.main()
