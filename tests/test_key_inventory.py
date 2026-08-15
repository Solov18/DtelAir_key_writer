import asyncio
import io
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import event
from starlette.requests import Request

from tests.postgres_test_case import PostgreSQLTestCase

import app.db as database
from app.repositories import key_repository, panel_repository
from app.routers import keys as keys_router
from app.routers.keys import key_assignment_update_route
from app.services.importer import import_keys_file
from app.services.writer import write_key_to_panels


class KeyInventoryTests(PostgreSQLTestCase):
    def setUp(self):
        super().setUp()

    def _create_key(self, key_type_id, number, hex_value):
        return key_repository.save_prepared_key(
            key_type_id,
            str(number),
            hex_value,
            "Тест",
        )

    @staticmethod
    def _write_request() -> Request:
        return Request(
            {
                "type": "http",
                "method": "POST",
                "path": "/message/write",
                "headers": [],
                "client": ("127.0.0.1", 50000),
                "session": {
                    "user": {
                        "login": "admin",
                        "full_name": "Администратор",
                        "role": "admin",
                    }
                },
            }
        )

    @staticmethod
    def _success_result():
        return {
            "ok": True,
            "written": True,
            "status": "SUCCESS",
            "response": "Ключ успешно записан",
            "message": "Ключ успешно записан",
        }

    @staticmethod
    def _create_panel(address: str, entrance: str, suffix: int) -> dict:
        mac = f"08:13:CD:20:00:{suffix:02X}"
        panel_repository.create_or_update_panel(address, entrance, mac=mac)
        with database.db() as conn:
            return dict(
                conn.execute("SELECT * FROM panels WHERE mac = ?", (mac,)).fetchone()
            )

    def test_number_is_unique_only_inside_type(self):
        blue_id = key_repository.create_key_type("Синий", "#168EE8")
        orange_id = key_repository.create_key_type("Оранжевый", "#FF982A")

        blue = self._create_key(blue_id, 1, "AAAAAA01")
        orange = self._create_key(orange_id, 1, "BBBBBB01")

        self.assertNotEqual(blue["id"], orange["id"])
        with self.assertRaisesRegex(ValueError, "уже сохранён HEX"):
            self._create_key(blue_id, 1, "CCCCCC01")

    def test_key_types_expose_last_and_next_numeric_number(self):
        blue_id = key_repository.create_key_type("Синий", "#168EE8")
        empty_id = key_repository.create_key_type("Новый тип", "#22B889")
        padded_id = key_repository.create_key_type("С ведущими нулями", "#9B72E8")

        self._create_key(blue_id, 456788, "AABB0001")
        self._create_key(blue_id, 456789, "AABB0002")
        padded_batch = key_repository.prepare_key_range(
            padded_id,
            "0009",
            2,
            "Тест",
        )
        self._create_key(padded_id, "0009", "AABB0003")
        self._create_key(padded_id, "0010", "AABB0004")

        key_types = {
            item["id"]: item
            for item in key_repository.get_key_types(include_archived=False)
        }

        self.assertEqual(key_types[blue_id]["last_number"], "456789")
        self.assertEqual(key_types[blue_id]["next_number"], "456790")
        self.assertEqual(
            [item["number"] for item in padded_batch["rows"]],
            ["0009", "0010"],
        )
        self.assertEqual(key_types[padded_id]["last_number"], "0010")
        self.assertEqual(key_types[padded_id]["next_number"], "0011")
        self.assertEqual(key_types[empty_id]["last_number"], "")
        self.assertEqual(key_types[empty_id]["next_number"], "")

    def test_missing_numbers_are_found_inside_selected_key_type(self):
        blue_id = key_repository.create_key_type("Синий", "#168EE8")
        orange_id = key_repository.create_key_type("Оранжевый", "#FF982A")

        self._create_key(blue_id, "0100", "AABB0100")
        self._create_key(blue_id, "0102", "AABB0102")
        self._create_key(blue_id, "0105", "AABB0105")
        self._create_key(orange_id, "0101", "CCDD0101")

        result = key_repository.get_missing_key_numbers(blue_id)

        self.assertEqual(result["start"], "0100")
        self.assertEqual(result["end"], "0105")
        self.assertEqual(result["numbers"], ["0101", "0103", "0104"])
        self.assertEqual(result["missing_count"], 3)
        self.assertEqual(
            [(item["start"], item["end"], item["count"]) for item in result["ranges"]],
            [("0101", "0101", 1), ("0103", "0104", 2)],
        )

    def test_missing_numbers_support_explicit_range_and_ignore_blank_hex(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")
        self._create_key(key_type_id, "100", "AABB0100")
        self._create_key(key_type_id, "102", "AABB0102")
        with database.db() as conn:
            conn.execute(
                """
                INSERT INTO keys(key_type_id, number, hex_value, key_type, status)
                VALUES (?, '101', '', 'Синий', 'free')
                """,
                (key_type_id,),
            )

        result = key_repository.get_missing_key_numbers(
            key_type_id,
            "099",
            "103",
        )

        self.assertEqual(result["numbers"], ["99", "101", "103"])
        self.assertEqual(result["missing_count"], 3)

    def test_missing_numbers_compare_numerically_and_keep_type_display_format(self):
        blue_id = key_repository.create_key_type("Синий без дополнения", "#168EE8")
        orange_id = key_repository.create_key_type("Оранжевый шестизначный", "#FF982A")

        for number in ("1", "2", "4", "5"):
            self._create_key(blue_id, number, f"AA{int(number):06d}")
        for number in ("000001", "000002", "000004", "000005"):
            self._create_key(orange_id, number, f"BB{int(number):06d}")

        blue = key_repository.get_missing_key_numbers(blue_id, "1", "5")
        orange = key_repository.get_missing_key_numbers(orange_id, "1", "5")

        self.assertEqual(blue["start"], "1")
        self.assertEqual(blue["end"], "5")
        self.assertEqual(blue["numbers"], ["3"])
        self.assertEqual(blue["ranges"], [{"start": "3", "end": "3", "count": 1}])
        self.assertEqual(orange["start"], "000001")
        self.assertEqual(orange["end"], "000005")
        self.assertEqual(orange["numbers"], ["000003"])
        self.assertEqual(
            orange["ranges"],
            [{"start": "000003", "end": "000003", "count": 1}],
        )

    def test_large_unpadded_number_does_not_pad_small_missing_range(self):
        blue_id = key_repository.create_key_type("Синий переменной длины", "#168EE8")
        for number in ("1", "2", "4", "5", "40630"):
            self._create_key(blue_id, number, f"CC{int(number):06d}")

        result = key_repository.get_missing_key_numbers(blue_id, "1", "5")

        self.assertEqual(result["start"], "1")
        self.assertEqual(result["end"], "5")
        self.assertEqual(result["numbers"], ["3"])

    def test_scanner_rejects_duplicate_hex(self):
        key_type_id = key_repository.create_key_type("Стикер", "#9B72E8")
        batch = key_repository.prepare_key_range(key_type_id, 10, 2, "Тест")
        first, second = batch["rows"]

        key_repository.save_prepared_key(
            key_type_id,
            first["number"],
            "363FFAD7",
            "Тест",
        )

        with self.assertRaisesRegex(ValueError, "уже принадлежит"):
            key_repository.save_prepared_key(
                key_type_id,
                second["number"],
                "36:3F:FA:D7",
                "Тест",
            )

    def test_prepare_skips_filled_numbers_and_protects_saved_hex(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")
        original = key_repository.prepare_key_range(key_type_id, 10, 2, "Тест")
        first, second = original["rows"]
        saved_first = key_repository.save_prepared_key(
            key_type_id,
            first["number"],
            "AABBCCDD",
            "Тест",
        )

        repeated = key_repository.prepare_key_range(key_type_id, 10, 2, "Тест")

        self.assertEqual(repeated["created"], 0)
        self.assertEqual(repeated["filled_existing"], 1)
        self.assertEqual(repeated["resumed"], 0)
        self.assertEqual([row["number"] for row in repeated["rows"]], [second["number"]])

        with self.assertRaisesRegex(ValueError, "уже сохранён HEX"):
            key_repository.save_prepared_key(
                key_type_id,
                first["number"],
                "11223344",
                "Тест",
            )

        corrected = key_repository.save_prepared_key(
            key_type_id,
            first["number"],
            "11223344",
            "Тест",
            allow_replace=True,
        )
        self.assertEqual(corrected["id"], saved_first["id"])
        self.assertEqual(corrected["hex_value"], "11223344")

    def test_preparation_does_not_create_keys_without_hex(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")

        batch = key_repository.prepare_key_range(key_type_id, 500001, 3, "Тест")

        with database.db() as conn:
            keys_before_scan = conn.execute("SELECT COUNT(*) FROM keys").fetchone()[0]
        self.assertEqual(keys_before_scan, 0)
        self.assertEqual(
            [row["number"] for row in batch["rows"]],
            ["500001", "500002", "500003"],
        )

        saved = self._create_key(key_type_id, 500001, "ABCD0001")
        with database.db() as conn:
            stored = conn.execute(
                "SELECT number, hex_value FROM keys WHERE id = ?",
                (saved["id"],),
            ).fetchone()
        self.assertEqual(tuple(stored), ("500001", "ABCD0001"))

    def test_legacy_blank_rows_are_hidden_and_reused(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")
        with database.db() as conn:
            blank_id = conn.execute(
                """
                INSERT INTO keys(key_type_id, number, hex_value, key_type, status)
                VALUES (?, '500099', '', 'Синий', 'free')
                """,
                (key_type_id,),
            ).lastrowid

        self.assertEqual(key_repository.get_keys_page()["total"], 0)
        self.assertEqual(key_repository.get_key_statistics()["total"], 0)
        key_type = next(
            item for item in key_repository.get_key_types() if item["id"] == key_type_id
        )
        self.assertEqual(key_type["keys_count"], 0)

        saved = self._create_key(key_type_id, 500099, "ABCD0099")

        self.assertEqual(saved["id"], blank_id)
        self.assertEqual(key_repository.get_keys_page()["total"], 1)
        self.assertEqual(key_repository.get_key_statistics()["total"], 1)

    def test_existing_key_cannot_be_cleared_or_blank_key_assigned(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")
        saved = self._create_key(key_type_id, 1523, "ABCD1523")

        with self.assertRaisesRegex(ValueError, "нельзя сохранить без HEX"):
            key_repository.update_key(
                saved["id"],
                key_type_id,
                "1523",
                "",
                "",
            )

        with database.db() as conn:
            blank_id = conn.execute(
                """
                INSERT INTO keys(key_type_id, number, hex_value, key_type, status)
                VALUES (?, '1524', '', 'Синий', 'free')
                """,
                (key_type_id,),
            ).lastrowid

        with self.assertRaisesRegex(ValueError, "без HEX нельзя назначить"):
            key_repository.set_key_assignment(blank_id, "resident", apartment="15")

    def test_assignment_updates_status_and_release_keeps_history(self):
        key_type_id = key_repository.create_key_type("Премиум", "#E8B630")
        key = self._create_key(key_type_id, 77, "AABBCCDD")

        key_repository.set_key_assignment(
            key["id"],
            "resident",
            address="Тепличная 63",
            apartment="15",
            assigned_by="Оператор",
        )
        assigned = key_repository.get_key(key["id"])

        self.assertEqual(assigned["status"], "issued_resident")
        self.assertEqual(assigned["assignment_text"], "Тепличная 63 / кв. 15")

        key_repository.release_key(key["id"], "Возвращён")
        released = key_repository.get_key(key["id"])
        assignments = key_repository.get_key_assignments(key["id"])

        self.assertEqual(released["status"], "free")
        self.assertIsNone(released["assignment_type"])
        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0]["active"], 0)

    def test_assignment_can_be_edited_without_panel_requests_and_keeps_history(self):
        key_type_id = key_repository.create_key_type("Оранжевый", "#FF982A")
        key = self._create_key(key_type_id, "003602", "A0F0BD52")
        with database.db() as conn:
            conn.execute(
                """
                INSERT INTO panels(address, entrance, name, mac, enabled)
                VALUES ('ул. Голубые Дали 80', 'подъезд 1', 'Голубые Дали 80',
                        '08:13:CD:80:00:01', 1)
                """
            )
        key_repository.set_key_assignment(
            key["id"],
            "resident",
            address="ул. Голубые Дали 80",
            assigned_by="Импорт",
            note="Перенос старой базы",
        )
        request = Request(
            {
                "type": "http",
                "method": "POST",
                "path": f"/keys/{key['id']}/assignment",
                "headers": [],
                "session": {
                    "user": {
                        "login": "admin",
                        "full_name": "Соловьёв Евгений",
                        "role": "admin",
                    }
                },
                "client": ("127.0.0.1", 12345),
            }
        )

        with patch("app.services.writer.crm_add_key") as crm_add_key:
            response = key_assignment_update_route(
                request,
                key["id"],
                assignment_type="uk",
                address="ул. Голубые Дали 80",
                apartment="",
                owner_name="УК Голубые Дали",
                reason="Исправление владельца при переносе старой базы",
            )

        crm_add_key.assert_not_called()
        self.assertEqual(response.status_code, 303)
        updated = key_repository.get_key(key["id"])
        assignments = key_repository.get_key_assignments(key["id"])
        self.assertEqual(updated["status"], "assigned_uk")
        self.assertEqual(updated["number"], "003602")
        self.assertEqual(updated["hex_value"], "A0F0BD52")
        self.assertEqual(len(assignments), 2)
        self.assertEqual(assignments[0]["assignment_type"], "uk")
        self.assertEqual(assignments[0]["address"], "ул. Голубые Дали 80")
        self.assertEqual(assignments[0]["apartment"], "")
        self.assertEqual(assignments[0]["note"], "УК Голубые Дали")
        self.assertEqual(assignments[0]["active"], 1)
        self.assertEqual(assignments[1]["assignment_type"], "resident")
        self.assertEqual(assignments[1]["active"], 0)
        with database.db() as conn:
            operation = dict(
                conn.execute(
                    """
                    SELECT action, details, address, apartment, username
                    FROM operation_log
                    WHERE key_id = ? AND action = 'key_assignment_update'
                    """,
                    (key["id"],),
                ).fetchone()
            )
        self.assertIn("Жилец, ул. Голубые Дали 80, квартира не указана", operation["details"])
        self.assertIn("УК, ул. Голубые Дали 80, квартира не указана, УК Голубые Дали", operation["details"])
        self.assertIn("Исправление владельца", operation["details"])
        self.assertEqual(operation["address"], "ул. Голубые Дали 80")
        self.assertEqual(operation["apartment"], "")
        self.assertEqual(operation["username"], "admin")

    def test_assignment_edit_rejects_unknown_address_without_changing_current_row(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")
        key = self._create_key(key_type_id, "42", "AABB0042")
        key_repository.set_key_assignment(
            key["id"],
            "resident",
            address="ул. Существующая 1",
            apartment="7",
        )

        with self.assertRaisesRegex(ValueError, "Адрес не найден"):
            key_repository.update_key_assignment(
                key["id"],
                "other",
                address="ул. Несуществующая 99",
                owner_name="Техническая служба",
                reason="Проверка",
            )

        assignments = key_repository.get_key_assignments(key["id"])
        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0]["active"], 1)
        self.assertEqual(assignments[0]["assignment_type"], "resident")

    def test_excel_import_rejects_rows_without_hex(self):
        workbook = Workbook()
        blue = workbook.active
        blue.title = "Синий"
        blue.append(["№", "HEX", "Комментарий"])
        blue.append([1, "363FFAD7", "Первый"])
        blue.append([2, "", "Заготовка"])
        orange = workbook.create_sheet("Оранжевый")
        orange.append(["Number", "Код"])
        orange.append([1, "3644D427"])
        content = io.BytesIO()
        workbook.save(content)

        first_report = import_keys_file("keys.xlsx", content.getvalue(), "Тест")

        self.assertEqual(first_report["created_types"], 2)
        self.assertEqual(first_report["added"], 2)
        self.assertEqual(first_report["errors"], 1)
        self.assertTrue(any("не указан HEX" in item for item in first_report["error_details"]))

        blue[3][1].value = "11223344"
        updated_content = io.BytesIO()
        workbook.save(updated_content)
        second_report = import_keys_file(
            "keys.xlsx",
            updated_content.getvalue(),
            "Тест",
        )

        self.assertEqual(second_report["added"], 1)
        self.assertGreaterEqual(second_report["duplicates"], 2)

    def test_excel_import_preserves_leading_zeroes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Оранжевый"
        sheet.append(["Номер", "HEX"])
        sheet.append(["000050", "AA000050"])
        sheet.append(["50", "AA000051"])
        content = io.BytesIO()
        workbook.save(content)

        report = import_keys_file("keys.xlsx", content.getvalue(), "Тест")

        self.assertEqual(report["added"], 2)
        with database.db() as conn:
            numbers = [
                row[0]
                for row in conn.execute(
                    "SELECT number FROM keys ORDER BY id"
                ).fetchall()
            ]
        self.assertEqual(numbers, ["000050", "50"])

    def test_large_excel_import_uses_one_key_snapshot_and_one_batch_insert(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Синий"
        sheet.append(["№", "HEX", "Комментарий"])
        for number in range(1, 41):
            sheet.append([number, f"AA{number:06X}", "Пакетный импорт"])
        content = io.BytesIO()
        workbook.save(content)

        statements: list[tuple[str, bool]] = []

        def capture(_conn, _cursor, statement, _parameters, _context, executemany):
            statements.append((" ".join(statement.lower().split()), bool(executemany)))

        engine = database.get_engine()
        event.listen(engine, "before_cursor_execute", capture)
        try:
            report = import_keys_file("keys.xlsx", content.getvalue(), "Тест")
        finally:
            event.remove(engine, "before_cursor_execute", capture)

        key_snapshots = [
            statement
            for statement, _many in statements
            if statement.startswith("select id, key_type_id, number, hex_value")
            and " from keys" in statement
        ]
        key_inserts = [
            many
            for statement, many in statements
            if statement.startswith("insert into keys(")
        ]
        self.assertEqual(report["added"], 40)
        self.assertEqual(len(key_snapshots), 1)
        self.assertEqual(key_inserts, [True])

    def test_key_export_has_download_headers_and_valid_empty_workbook(self):
        with patch.object(keys_router, "get_all_keys_for_export", return_value=[]):
            response = keys_router.keys_export()

        async def read_response() -> bytes:
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(chunk)
            return b"".join(chunks)

        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(
            response.headers["content-disposition"],
            'attachment; filename="keys_export.xlsx"',
        )
        workbook = load_workbook(
            io.BytesIO(asyncio.run(read_response())),
            read_only=True,
        )
        self.assertEqual(workbook.sheetnames, ["Ключи"])
        self.assertEqual(workbook["Ключи"]["A1"].value, "Номер")

    def test_key_export_preserves_padded_number_as_excel_text(self):
        exported = [{
            "type_name": "Оранжевый",
            "number": "000050",
            "hex_value": "AA000050",
            "status": "free",
            "note": "",
            "created_at": "2026-08-14 10:00:00",
            "created_by": "Тест",
        }]
        with patch.object(keys_router, "get_all_keys_for_export", return_value=exported):
            response = keys_router.keys_export()

        async def read_response() -> bytes:
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(chunk)
            return b"".join(chunks)

        workbook = load_workbook(io.BytesIO(asyncio.run(read_response())))
        cell = workbook["Оранжевый"]["A2"]
        self.assertEqual(cell.value, "000050")
        self.assertEqual(cell.data_type, "s")
        self.assertEqual(cell.number_format, "@")

    def test_problem_key_is_not_sent_to_crm(self):
        key_type_id = key_repository.create_key_type("Детский", "#5CC878")
        key = self._create_key(key_type_id, 15, "A1B2C3D4")
        key_repository.set_key_status(key["id"], "blocked", "Проверка")
        blocked_key = key_repository.get_key(key["id"])

        with patch("app.services.writer.crm_add_key") as crm_add_key:
            results = write_key_to_panels(
                "resident_manual",
                blocked_key,
                [{"id": 1, "name": "Вход", "mac": "08:55:CD:00:00:01"}],
                flat_num="15",
                address="Тепличная 63",
            )

        crm_add_key.assert_not_called()
        self.assertEqual(results[0]["status"], "KEY_UNAVAILABLE")
        self.assertFalse(results[0]["written"])

    def test_generic_uk_assignment_replaces_employee_and_can_be_released(self):
        key_type_id = key_repository.create_key_type("Служебный", "#159ED9")
        key = self._create_key(key_type_id, 91, "A1B2C391")

        with database.db() as conn:
            employee_id = conn.execute(
                "INSERT INTO employees(full_name) VALUES ('Иванов Иван')"
            ).lastrowid
            group_id = conn.execute(
                "INSERT INTO uk_groups(name) VALUES ('УК Тест')"
            ).lastrowid
            conn.execute(
                """
                INSERT INTO employee_keys(employee_id, key_id, status)
                VALUES (?, ?, 'active')
                """,
                (employee_id, key["id"]),
            )

        key_repository.set_key_assignment(
            key["id"],
            "uk",
            uk_group_id=group_id,
            assigned_by="Тест",
        )

        with database.db() as conn:
            employee_status = conn.execute(
                "SELECT status FROM employee_keys WHERE key_id = ?",
                (key["id"],),
            ).fetchone()[0]
            uk_assignments = conn.execute(
                """
                SELECT COUNT(*)
                FROM key_assignments
                WHERE key_id = ?
                  AND assignment_type = 'uk'
                  AND active = 1
                """,
                (key["id"],),
            ).fetchone()[0]

        self.assertEqual(employee_status, "replaced")
        self.assertEqual(uk_assignments, 1)
        self.assertEqual(key_repository.get_key(key["id"])["status"], "assigned_uk")

        key_repository.release_key(key["id"], "Возвращён")

        with database.db() as conn:
            uk_assignments_after_release = conn.execute(
                """
                SELECT COUNT(*)
                FROM key_assignments
                WHERE key_id = ?
                  AND assignment_type = 'uk'
                  AND active = 1
                """,
                (key["id"],),
            ).fetchone()[0]

        self.assertEqual(uk_assignments_after_release, 0)
        self.assertEqual(key_repository.get_key(key["id"])["status"], "free")

    def test_released_assignment_is_not_returned_as_current_search_result(self):
        key_type_id = key_repository.create_key_type("Синий", "#168EE8")
        key = self._create_key(key_type_id, 1523, "363FFAD7")

        with patch(
            "app.services.writer.crm_add_key",
            return_value={
                "ok": True,
                "written": True,
                "status": "SUCCESS",
                "response": "Ключ успешно записан",
                "message": "Ключ успешно записан",
            },
        ):
            write_key_to_panels(
                "resident_manual",
                key,
                [
                    {
                        "id": 42,
                        "name": "Подъезд 1",
                        "mac": "08:55:CD:00:00:01",
                        "address": "Тепличная 63",
                    }
                ],
                flat_num="15",
                address="Тепличная 63",
                assignment_type="resident",
            )

        key_repository.release_key(key["id"], "Перенесён")

        with database.db() as conn:
            panel_id = conn.execute(
                "SELECT panel_id FROM operation_log WHERE key_id = ?",
                (key["id"],),
            ).fetchone()[0]

        self.assertEqual(panel_id, 42)
        self.assertEqual(key_repository.get_keys_page(query="Тепличная 63")["total"], 0)
        self.assertEqual(key_repository.get_keys_page(query="кв 15")["total"], 0)

    def _create_assignment_search_fixture(self):
        blue_id = key_repository.create_key_type("Синий", "#168EE8")
        orange_id = key_repository.create_key_type("Оранжевый", "#FF982A")
        exact = self._create_key(blue_id, 7001, "AA000001")
        other_flat = self._create_key(blue_id, 7002, "AA000002")
        other_house = self._create_key(orange_id, 7003, "AA000003")
        misleading_number = self._create_key(blue_id, 32, "AA000032")
        uk_key = self._create_key(orange_id, 7004, "AA000004")
        employee_key = self._create_key(blue_id, 7005, "AA000005")

        key_repository.set_key_assignment(
            exact["id"],
            "resident",
            address="ул. Тепличная, д. 71/5",
            apartment="32",
            assigned_by="Тест",
        )
        key_repository.set_key_assignment(
            other_flat["id"],
            "resident",
            address="Тепличная 71 корпус 5",
            apartment="33",
            assigned_by="Тест",
        )
        key_repository.set_key_assignment(
            other_house["id"],
            "resident",
            address="Тепличная 71",
            apartment="32",
            assigned_by="Тест",
        )
        with database.db() as conn:
            uk_group_id = conn.execute(
                "INSERT INTO uk_groups(name) VALUES ('УК Малышева')"
            ).lastrowid
            employee_id = conn.execute(
                "INSERT INTO employees(full_name) VALUES ('Иванов Сергей Петрович')"
            ).lastrowid
        key_repository.set_key_assignment(
            uk_key["id"],
            "uk",
            address="ул. Малышева 7",
            uk_group_id=uk_group_id,
            assigned_by="Тест",
            note="УК Малышева",
        )
        key_repository.set_key_assignment(
            employee_key["id"],
            "employee",
            employee_id=employee_id,
            assigned_by="Тест",
        )
        return {
            "blue_id": blue_id,
            "orange_id": orange_id,
            "exact": exact,
            "other_flat": other_flat,
            "other_house": other_house,
            "misleading_number": misleading_number,
            "uk_key": uk_key,
            "employee_key": employee_key,
        }

    def test_key_register_searches_number_and_hex(self):
        fixture = self._create_assignment_search_fixture()

        by_number = key_repository.get_keys_page(query="7001")
        by_hex = key_repository.get_keys_page(query="aa:00:00:01")

        self.assertEqual([item["id"] for item in by_number["items"]], [fixture["exact"]["id"]])
        self.assertEqual([item["id"] for item in by_hex["items"]], [fixture["exact"]["id"]])

    def test_key_register_searches_exact_house(self):
        fixture = self._create_assignment_search_fixture()

        result = key_repository.get_keys_page(query="ул Тепличная дом 71/5")

        self.assertEqual(
            {item["id"] for item in result["items"]},
            {fixture["exact"]["id"], fixture["other_flat"]["id"]},
        )

    def test_key_register_requires_exact_apartment_with_address(self):
        fixture = self._create_assignment_search_fixture()

        result = key_repository.get_keys_page(query="Тепличная 71/5 квартира №32")

        self.assertEqual([item["id"] for item in result["items"]], [fixture["exact"]["id"]])
        self.assertEqual(result["items"][0]["assignment_address"], "ул. Тепличная, д. 71/5")
        self.assertEqual(result["items"][0]["assignment_apartment"], "32")

    def test_key_register_searches_only_by_apartment(self):
        fixture = self._create_assignment_search_fixture()

        result = key_repository.get_keys_page(query="кв. 32")

        self.assertEqual(
            {item["id"] for item in result["items"]},
            {fixture["exact"]["id"], fixture["other_house"]["id"]},
        )
        self.assertNotIn(fixture["misleading_number"]["id"], {item["id"] for item in result["items"]})

    def test_key_register_searches_assignment_owner_and_uk(self):
        fixture = self._create_assignment_search_fixture()

        result = key_repository.get_keys_page(query="УК Малышева")

        self.assertEqual([item["id"] for item in result["items"]], [fixture["uk_key"]["id"]])
        self.assertEqual(result["items"][0]["assignment_type_name"], "УК")
        self.assertEqual(result["items"][0]["uk_name"], "УК Малышева")

        employee_result = key_repository.get_keys_page(query="Иванов Сергей")
        self.assertEqual(
            [item["id"] for item in employee_result["items"]],
            [fixture["employee_key"]["id"]],
        )
        self.assertEqual(
            employee_result["items"][0]["assignment_text"],
            "Иванов Сергей Петрович",
        )

    def test_assignment_search_combines_with_type_and_status_filters(self):
        fixture = self._create_assignment_search_fixture()

        blue = key_repository.get_keys_page(
            query="Тепличная 71/5",
            key_type_id=fixture["blue_id"],
        )
        orange = key_repository.get_keys_page(
            query="Тепличная 71/5",
            key_type_id=fixture["orange_id"],
        )
        issued = key_repository.get_keys_page(
            query="Тепличная 71/5 кв32",
            status="issued_resident",
        )
        free = key_repository.get_keys_page(
            query="Тепличная 71/5 кв32",
            availability="free",
        )

        self.assertEqual(blue["total"], 2)
        self.assertEqual(orange["total"], 0)
        self.assertEqual([item["id"] for item in issued["items"]], [fixture["exact"]["id"]])
        self.assertEqual(free["total"], 0)

    def test_free_key_is_written_and_receives_assignment(self):
        type_id = key_repository.create_key_type("Синий", "#168EE8")
        key = self._create_key(type_id, 8101, "CC008101")
        panel = self._create_panel("Тестовая 10", "Подъезд 1", 1)

        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ) as crm:
            results = write_key_to_panels(
                "message",
                key,
                [panel],
                flat_num="12",
                address="Тестовая 10",
                request=self._write_request(),
                assignment_type="resident",
                write_option="write_free_key",
            )

        crm.assert_called_once()
        self.assertEqual(results[0]["status"], "SUCCESS")
        current = key_repository.get_key_assignments(key["id"])[0]
        self.assertTrue(current["active"])
        self.assertEqual(current["address"], "Тестовая 10")
        self.assertEqual(current["apartment"], "12")

    def test_used_key_can_be_reassigned_without_removing_old_panels(self):
        type_id = key_repository.create_key_type("Оранжевый", "#FF982A")
        key = self._create_key(type_id, 8102, "CC008102")
        panel = self._create_panel("Новая 20", "Подъезд 2", 2)
        key_repository.set_key_assignment(
            key["id"], "resident", address="Старая 1", apartment="4"
        )

        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ):
            write_key_to_panels(
                "message",
                key_repository.get_key(key["id"]),
                [panel],
                flat_num="8",
                address="Новая 20",
                request=self._write_request(),
                assignment_type="resident",
                assignment_policy="replace",
                write_option="reassign_to_new_address",
                previous_assignment="Жилец, Старая 1, кв. 4",
            )

        assignments = key_repository.get_key_assignments(key["id"])
        self.assertEqual(len(assignments), 2)
        self.assertEqual(assignments[0]["address"], "Новая 20")
        self.assertTrue(assignments[0]["active"])
        self.assertFalse(assignments[1]["active"])

    def test_used_key_can_add_panels_without_changing_assignment(self):
        type_id = key_repository.create_key_type("Служебный", "#9B72E8")
        key = self._create_key(type_id, 8103, "CC008103")
        panel = self._create_panel("Другой дом 3", "Вход", 3)
        key_repository.set_key_assignment(
            key["id"], "uk", address="Дом УК 1", note="УК Тест"
        )

        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ):
            write_key_to_panels(
                "message",
                key_repository.get_key(key["id"]),
                [panel],
                flat_num="15",
                address="Другой дом 3",
                request=self._write_request(),
                assignment_policy="preserve",
                write_option="add_selected_panels",
                previous_assignment="УК, Дом УК 1, УК Тест",
            )

        assignments = key_repository.get_key_assignments(key["id"])
        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0]["assignment_type"], "uk")
        self.assertEqual(assignments[0]["address"], "Дом УК 1")
        context = key_repository.get_key_write_contexts([key["id"]])[key["id"]]
        self.assertEqual(context["panel_ids"], [panel["id"]])

    def test_key_already_on_one_selected_panel_writes_only_missing_panel(self):
        type_id = key_repository.create_key_type("Синий", "#168EE8")
        key = self._create_key(type_id, 8104, "CC008104")
        first = self._create_panel("Общий дом 4", "Подъезд 1", 4)
        second = self._create_panel("Общий дом 4", "Подъезд 2", 5)

        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ):
            write_key_to_panels(
                "message", key, [first], request=self._write_request()
            )
        context = key_repository.get_key_write_contexts([key["id"]])[key["id"]]
        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ) as crm:
            results = write_key_to_panels(
                "message",
                key_repository.get_key(key["id"]),
                [first, second],
                request=self._write_request(),
                assignment_policy="preserve",
                known_panel_ids=set(context["panel_ids"]),
                write_option="add_selected_panels",
            )

        crm.assert_called_once_with(second["mac"], key["hex_value"], "0", 1)
        self.assertEqual(
            [result["status"] for result in results],
            ["ALREADY_ON_PANEL", "SUCCESS"],
        )

    def test_key_on_all_selected_panels_sends_no_repeated_request(self):
        type_id = key_repository.create_key_type("Синий", "#168EE8")
        key = self._create_key(type_id, 8105, "CC008105")
        panel = self._create_panel("Общий дом 5", "Подъезд 1", 6)
        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ):
            write_key_to_panels(
                "message", key, [panel], request=self._write_request()
            )
        context = key_repository.get_key_write_contexts([key["id"]])[key["id"]]

        with patch("app.services.writer.crm_add_key") as crm:
            results = write_key_to_panels(
                "message",
                key_repository.get_key(key["id"]),
                [panel],
                request=self._write_request(),
                assignment_policy="preserve",
                known_panel_ids=set(context["panel_ids"]),
                write_option="add_selected_panels",
            )

        crm.assert_not_called()
        self.assertEqual(results[0]["status"], "ALREADY_ON_PANEL")

    def test_repeated_additional_panel_write_is_idempotent(self):
        type_id = key_repository.create_key_type("Уникальный", "#22B889")
        key = self._create_key(type_id, 8106, "CC008106")
        panel = self._create_panel("Повторная 6", "Вход", 7)
        with patch(
            "app.services.writer.crm_add_key",
            return_value=self._success_result(),
        ) as crm:
            first = write_key_to_panels(
                "message", key, [panel], request=self._write_request()
            )
            known = key_repository.get_key_write_contexts([key["id"]])[key["id"]]
            second = write_key_to_panels(
                "message",
                key_repository.get_key(key["id"]),
                [panel],
                request=self._write_request(),
                assignment_policy="preserve",
                known_panel_ids=set(known["panel_ids"]),
                write_option="add_selected_panels",
            )

        self.assertEqual(crm.call_count, 1)
        self.assertEqual(first[0]["status"], "SUCCESS")
        self.assertEqual(second[0]["status"], "ALREADY_ON_PANEL")


if __name__ == "__main__":
    unittest.main()
